#!/usr/bin/env python
"""
Script to import South African lottery data from Excel spreadsheet.
This script purges existing data and imports new data from the provided Excel file.
"""

import os
import sys
import logging
import json
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime, timedelta
from flask import current_app
from models import db, LotteryResult, Screenshot

# Set up logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def parse_date(date_str):
    """Parse date from string format to datetime object"""
    try:
        # Use our dedicated date parsing module
        from excel_date_utils import parse_excel_date
        
        # Handle NaN values
        if pd.isna(date_str):
            return None
            
        # Use the utility function for consistent date parsing
        result = parse_excel_date(date_str)
        
        # Log the result for debugging
        logger.info(f"Using excel_date_utils to parse '{date_str}' → {result}")
        
        return result
                
        # Try various date formats in order of likelihood
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        # Last resort - try to handle Excel date serial numbers
        try:
            if date_str.isdigit():
                # Excel date serial number conversion
                return datetime(1899, 12, 30) + timedelta(days=int(date_str))
        except:
            pass
                
        # If all formats fail, log the error but don't immediately fail
        # Try to extract any date-like pattern from the string
        import re
        date_patterns = [
            r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})',  # YYYY-MM-DD
            r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})'   # DD-MM-YYYY or MM-DD-YYYY
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, date_str)
            if match:
                try:
                    groups = match.groups()
                    if len(groups[0]) == 4:  # YYYY-MM-DD
                        return datetime(int(groups[0]), int(groups[1]), int(groups[2]))
                    else:  # DD-MM-YYYY or MM-DD-YYYY (assume DD-MM-YYYY)
                        return datetime(int(groups[2]), int(groups[1]), int(groups[0]))
                except:
                    continue
                    
        # If we still can't parse it, raise error
        logger.error(f"All date parsing attempts failed for: {date_str}")
        raise ValueError(f"Couldn't parse date: {date_str}")
    except Exception as e:
        logger.error(f"Error parsing date {date_str}: {str(e)}")
        return None

def parse_numbers(numbers_str):
    """Parse numbers from string format to list of integers"""
    try:
        if pd.isna(numbers_str):
            return []
            
        # Handle different formats
        if isinstance(numbers_str, str):
            # If numbers are comma-separated
            if ',' in numbers_str:
                numbers = [int(num.strip()) for num in numbers_str.split(',') if num.strip().isdigit()]
                if numbers:
                    return numbers
                    
            # If numbers are space-separated or have other delimiters
            clean_str = ''.join(c if c.isdigit() or c.isspace() else ' ' for c in numbers_str)
            numbers = [int(num) for num in clean_str.split() if num.strip().isdigit()]
            return numbers
        elif isinstance(numbers_str, (int, float)):
            # Single number
            return [int(numbers_str)]
        else:
            return []
    except Exception as e:
        logger.error(f"Error parsing numbers {numbers_str}: {str(e)}")
        return []

def parse_divisions(divisions_data):
    """Parse divisions data from various formats to a structured dictionary"""
    if pd.isna(divisions_data):
        return {}
        
    divisions = {}
    try:
        # If it's already a string that looks like JSON
        if isinstance(divisions_data, str) and ('{' in divisions_data or ':' in divisions_data):
            try:
                # Try to parse as JSON if it looks like a dictionary
                if '{' in divisions_data:
                    return json.loads(divisions_data)
                
                # Try to parse from semicolon-separated format
                # Division 1: 0 winners, R0.00; Division 2: 5 winners, R100,563.40
                if ';' in divisions_data:
                    division_parts = divisions_data.split(';')
                    for part in division_parts:
                        if ':' in part:
                            div_name, div_data = part.split(':', 1)
                            div_name = div_name.strip()
                            div_data = div_data.strip()
                            if ',' in div_data:
                                winners_part, prize_part = div_data.split(',', 1)
                                # Clean up winners part to extract just the number
                                winners = ''.join(c for c in winners_part if c.isdigit())
                                # Keep the prize as is with R and commas
                                prize = prize_part.strip()
                                divisions[div_name] = {
                                    "winners": winners,
                                    "prize": prize
                                }
            except Exception as e:
                logger.warning(f"Error parsing divisions string: {str(e)}")
                
        # For more complex division data that might be in columns
        # We'll handle this in the main import function
                
    except Exception as e:
        logger.error(f"Error parsing divisions: {str(e)}")
        
    return divisions

def standardize_lottery_type(lottery_type):
    """Standardize lottery type names for consistency"""
    if pd.isna(lottery_type):
        return ""
        
    # Convert to string, strip whitespace, and lowercase for comparison
    lt = str(lottery_type).strip().lower()
    
    # Enhanced logging for troubleshooting
    logger.debug(f"Standardizing lottery type: '{lottery_type}' (lowercase: '{lt}')")
    
    # Normalize common variations with improved pattern matching - prioritize "Lottery" over "Lotto"
    if ('lottery plus 1' in lt or 'lotto plus 1' in lt or 'lottery+1' in lt or 'lotto+1' in lt or
        'lottery + 1' in lt or 'lotto + 1' in lt or 'lottery plus1' in lt or 'lotto plus1' in lt or
        lt == 'lottery plus 1' or lt == 'lotto plus 1' or 'lotteryplus1' in lt or 'lottoplus1' in lt):
        return 'Lottery Plus 1'
    elif ('lottery plus 2' in lt or 'lotto plus 2' in lt or 'lottery+2' in lt or 'lotto+2' in lt or
          'lottery + 2' in lt or 'lotto + 2' in lt or 'lottery plus2' in lt or 'lotto plus2' in lt or
          lt == 'lottery plus 2' or lt == 'lotto plus 2' or 'lotteryplus2' in lt or 'lottoplus2' in lt):
        return 'Lottery Plus 2'
    elif ('powerball plus' in lt or 'powerball+' in lt or 'power ball plus' in lt or 
          'powerballplus' in lt or lt == 'powerball plus'):
        return 'Powerball Plus'
    elif ('powerball' in lt or 'power ball' in lt or lt == 'powerball'):
        # Only match if it's not "powerball plus"
        if 'plus' not in lt:
            return 'Powerball'
    elif ('daily lottery' in lt or 'dailylottery' in lt or 'daily lotto' in lt or 'dailylotto' in lt or
          lt == 'daily lottery' or lt == 'dailylottery' or lt == 'daily lotto' or lt == 'dailylotto'):
        return 'Daily Lottery'
    # Main Lottery/Lotto - match only if it contains "lottery" or "lotto" but NOT "plus" or "daily"
    # Prioritize searching for "lottery" first
    elif ('lottery' in lt or 'lotto' in lt) and 'plus' not in lt and 'daily' not in lt:
        # Log when we standardize to "Lottery" for debugging
        if lottery_type.upper() == 'LOTTERY' or lottery_type.upper() == 'LOTTO':
            logger.info(f"Standardizing '{lottery_type}' to 'Lottery'")
        return 'Lottery'
    
    # If no match, log a warning and return original with proper capitalization
    logger.warning(f"No standard match found for lottery type: '{lottery_type}'")
    return str(lottery_type).strip()

def import_excel_data(excel_file, flask_app=None):
    """
    Import lottery data from Excel spreadsheet.
    
    Args:
        excel_file (str): Path to Excel file
        flask_app: Flask app object for context (optional)
    """
    try:
        # Use current_app or provided app
        ctx = flask_app.app_context() if flask_app else current_app.app_context()
        
        with ctx:
            logger.info(f"Starting import from {excel_file}...")
            # Initialize counters and tracking lists at the beginning of the function
            imported_count = 0
            updated_count = 0
            error_count = 0
            imported_records = []
            
            # Read Excel file - try multiple sheets if needed
            try:
                # First try reading with default sheet
                df = pd.read_excel(excel_file)
                
                # FIX for Row 1 (index 0) - Special direct processing of the FIRST ROW
                # This ensures the data in row 1 is NEVER missed, regardless of any skipping
                if not df.empty:
                    # Get raw row 1 data (index 0)
                    logger.warning("🔍 DIRECT ROW 1 ACCESS: Processing first Excel row directly")
                    row_1 = df.iloc[0]
                    
                    # Extract basic data directly
                    game_name = row_1['Game Name'] if 'Game Name' in row_1 else None
                    
                    if game_name is not None and not pd.isna(game_name):
                        logger.warning(f"🎯 DIRECT ROW 1: Found game {game_name}")
                        draw_number = str(row_1['Draw Number']).strip() if 'Draw Number' in row_1 else None
                        draw_date = row_1['Draw Date'] if 'Draw Date' in row_1 else None
                        winning_numbers = row_1['Winning Numbers (Numerical)'] if 'Winning Numbers (Numerical)' in row_1 else None
                        bonus_ball = row_1['Bonus Ball'] if 'Bonus Ball' in row_1 else None
                        
                        # Standardize lottery type directly
                        lottery_type = str(game_name).strip()
                        if lottery_type.upper() == 'LOTTO' or lottery_type.upper() == 'LOTTERY':
                            lottery_type = 'Lottery'
                        elif 'LOTTERY PLUS 1' in lottery_type.upper() or 'LOTTO PLUS 1' in lottery_type.upper():
                            lottery_type = 'Lottery Plus 1'
                        elif 'LOTTERY PLUS 2' in lottery_type.upper() or 'LOTTO PLUS 2' in lottery_type.upper():
                            lottery_type = 'Lottery Plus 2'
                        elif 'POWERBALL PLUS' in lottery_type.upper():
                            lottery_type = 'Powerball Plus'
                        elif 'POWERBALL' in lottery_type.upper():
                            lottery_type = 'Powerball'
                        elif 'DAILY LOTTERY' in lottery_type.upper() or 'DAILY LOTTO' in lottery_type.upper():
                            lottery_type = 'Daily Lottery'
                            
                        # Process numbers directly
                        if winning_numbers is not None and not pd.isna(winning_numbers):
                            if isinstance(winning_numbers, str):
                                numbers = [int(num.strip()) for num in winning_numbers.split() if num.strip().isdigit()]
                            elif isinstance(winning_numbers, (int, float)):
                                numbers = [int(winning_numbers)]
                            else:
                                numbers = []
                                
                            # Process bonus ball directly
                            if bonus_ball is not None and not pd.isna(bonus_ball):
                                if isinstance(bonus_ball, (int, float)):
                                    bonus_numbers = [int(bonus_ball)]
                                elif isinstance(bonus_ball, str):
                                    bonus_numbers = [int(num.strip()) for num in bonus_ball.split() if num.strip().isdigit()]
                                else:
                                    bonus_numbers = []
                            else:
                                bonus_numbers = []
                                
                            # Build divisions directly
                            divisions = {}
                            for i in range(1, 9):
                                winners_col = f'Div {i} Winners'
                                prize_col = f'Div {i} Winnings'
                                
                                if winners_col in row_1 and prize_col in row_1 and not pd.isna(row_1[winners_col]) and not pd.isna(row_1[prize_col]):
                                    winners_value = row_1[winners_col]
                                    prize_value = row_1[prize_col]
                                    
                                    # Format winners
                                    if isinstance(winners_value, (int, float)):
                                        winners_value = str(int(winners_value))
                                    else:
                                        winners_value = str(winners_value)
                                        
                                    # Format prize
                                    if isinstance(prize_value, str):
                                        if not prize_value.startswith("R"):
                                            prize_value = f"R{prize_value}"
                                    else:
                                        prize_value = f"R{prize_value}"
                                        
                                    divisions[f"Division {i}"] = {
                                        "winners": winners_value,
                                        "prize": prize_value
                                    }
                            
                            # Check if this result already exists in the database
                            if lottery_type and draw_number and draw_date and numbers:
                                logger.warning(f"✓ ROW 1 DIRECT: Valid data for {lottery_type} Draw {draw_number}")
                                
                                existing = LotteryResult.query.filter_by(
                                    lottery_type=lottery_type,
                                    draw_number=draw_number
                                ).first()
                                
                                if existing:
                                    logger.warning(f"✓ ROW 1 DIRECT: Updating existing {lottery_type} Draw {draw_number}")
                                    # Update existing record
                                    existing.draw_date = draw_date
                                    existing.numbers = json.dumps(numbers)
                                    existing.bonus_numbers = json.dumps(bonus_numbers) if bonus_numbers else None
                                    existing.divisions = json.dumps(divisions) if divisions else None
                                    existing.source_url = "imported-from-excel"
                                    existing.ocr_provider = "manual-import"
                                    existing.ocr_model = "excel-spreadsheet"
                                    existing.ocr_timestamp = datetime.utcnow().isoformat()
                                    db.session.commit()
                                    imported_count += 1
                                else:
                                    logger.warning(f"✓ ROW 1 DIRECT: Creating new entry for {lottery_type} Draw {draw_number}")
                                    # Create new result
                                    new_result = LotteryResult(
                                        lottery_type=lottery_type,
                                        draw_number=draw_number,
                                        draw_date=draw_date,
                                        numbers=json.dumps(numbers),
                                        bonus_numbers=json.dumps(bonus_numbers) if bonus_numbers else None,
                                        divisions=json.dumps(divisions) if divisions else None,
                                        source_url="imported-from-excel",
                                        ocr_provider="manual-import",
                                        ocr_model="excel-spreadsheet",
                                        ocr_timestamp=datetime.utcnow().isoformat()
                                    )
                                    db.session.add(new_result)
                                    db.session.commit()
                                    imported_count += 1
                                
                                # Store this record in imported_records for later reference
                                imported_records.append({
                                    'lottery_type': lottery_type,
                                    'draw_number': draw_number,
                                    'draw_date': draw_date,
                                    'is_new': existing is None,
                                    'lottery_result_id': (existing.id if existing else 
                                        LotteryResult.query.filter_by(lottery_type=lottery_type, draw_number=draw_number).first().id)
                                })
                                
                                logger.warning(f"✓ ROW 1 DIRECT: Successfully processed first Excel row")
                            else:
                                logger.warning("✗ ROW 1 DIRECT: Missing essential data")
                    else:
                        logger.warning("✗ DIRECT ROW 1: No valid Game Name found in first row")
                
                # If we got an empty dataframe or missing expected columns, try reading all sheets
                if df.empty or not any(col for col in df.columns if 'lottery' in str(col).lower() or 'lotto' in str(col).lower() or 'draw' in str(col).lower()):
                    logger.info("Initial sheet appears empty or missing lottery data. Checking all sheets...")
                    xlsx = pd.ExcelFile(excel_file)
                    
                    # Try each sheet until we find one with lottery data
                    for sheet_name in xlsx.sheet_names:
                        logger.info(f"Trying sheet: {sheet_name}")
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        # Check if this sheet has lottery data
                        if any(col for col in df.columns if 'lottery' in str(col).lower() or 'lotto' in str(col).lower() or 'draw' in str(col).lower() or 'game' in str(col).lower()):
                            logger.info(f"Found lottery data in sheet: {sheet_name}")
                            break
            except Exception as e:
                logger.error(f"Error reading Excel file: {str(e)}")
                return False
            
            # Replace NaN values with None for cleaner processing
            df = df.replace({np.nan: None})
            
            # Log column headers for debugging
            logger.info(f"Excel columns: {', '.join(str(col) for col in df.columns)}")
            
            # Map standard column names to actual columns in the spreadsheet
            column_mapping = {}
            
            # Check for common column name patterns in Snap Lottery format
            game_name_variants = ['game name', 'game type', 'lottery type', 'lottery game', 'lottery', 'lotto type']
            draw_number_variants = ['draw number', 'draw no', 'draw id', 'number']
            draw_date_variants = ['draw date', 'game date', 'date']
            numbers_variants = ['winning numbers', 'main numbers', 'numbers']
            bonus_variants = ['bonus ball', 'bonus number', 'powerball']
            
            for col in df.columns:
                col_str = str(col)
                col_lower = col_str.lower()
                
                # Map standard fields to actual column names with more flexible matching
                if any(variant in col_lower for variant in game_name_variants):
                    column_mapping['lottery_type'] = col
                
                # Special case for "Game Name" which is the most common in our templates
                if col_str == "Game Name":
                    column_mapping['lottery_type'] = col
                
                if any(variant in col_lower for variant in draw_number_variants):
                    column_mapping['draw_number'] = col
                
                if any(variant in col_lower for variant in draw_date_variants):
                    # If we have multiple date columns, prefer "Draw Date" over others
                    if 'draw_date' not in column_mapping or col_str == "Draw Date":
                        column_mapping['draw_date'] = col
                
                if any(variant in col_lower for variant in numbers_variants) or 'numerical' in col_lower:
                    column_mapping['numbers'] = col
                
                if any(variant in col_lower for variant in bonus_variants):
                    column_mapping['bonus_numbers'] = col
                
                # Division data handling - look for any columns related to divisions/winners/prizes
                if 'div' in col_lower or 'division' in col_lower or 'winners' in col_lower or 'prize' in col_lower or 'winnings' in col_lower:
                    # Note multiple division columns for later
                    if 'divisions' not in column_mapping:
                        column_mapping['divisions'] = []
                    column_mapping['divisions'].append(col)
            
            logger.info(f"Column mapping: {column_mapping}")
            
            # Count records before import
            initial_count = LotteryResult.query.count()
            logger.info(f"Database has {initial_count} records before import")
            
            # Using errors_count instead of error_count for consistency
            errors_count = error_count
            
            # CRITICAL FIX: Skip the header row before processing like in import_snap_lotto_data.py
            logger.warning("⚠️ CRITICAL: Using iloc[1:] to skip only the header row and prevent data loss")
            logger.warning("⚠️ This ensures row 2 in the Excel file (idx 1) is properly processed")
            # Skip only the first row (header) and reset the index
            df = df.iloc[1:].reset_index(drop=True)
            
            # Process each row
            for idx, row in df.iterrows():
                try:
                    # LOG ROW DETAILS for debugging - especially for row idx=0 (Excel row 2)
                    if idx == 0:
                        logger.warning(f"⚠️ Processing CRITICAL Row 2 (idx=0): {row.to_dict()}")
                    
                    # RELAXED EMPTINESS CHECK: Only skip if row appears completely empty (75%+ of values are None/NaN)
                    # This is a more forgiving check that will process partially filled rows
                    if row.isna().sum() > len(row) * 0.75:
                        logger.warning(f"Skipping row {idx+2} as it's very empty ({row.isna().sum()}/{len(row)} empty values)")
                        continue
                    
                    # Extract data using our column mapping
                    lottery_type = standardize_lottery_type(row.get(column_mapping.get('lottery_type', ''), None))
                    
                    # Enhanced draw number extraction with better normalization
                    raw_draw_number = str(row.get(column_mapping.get('draw_number', ''), '')).strip()
                    
                    # Try to use the normalize_draw_number function from data_aggregator
                    try:
                        # Import here to avoid circular imports
                        from data_aggregator import normalize_draw_number
                        draw_number = normalize_draw_number(raw_draw_number)
                        # Log the normalization if it changed
                        if raw_draw_number != draw_number:
                            logger.info(f"Normalized draw number from '{raw_draw_number}' to '{draw_number}'")
                    except ImportError:
                        # Fallback to basic cleaning if importing fails
                        draw_number = raw_draw_number
                        # Remove "Draw" prefix if present
                        draw_number = draw_number.replace('Draw', '').replace('DRAW', '').strip()
                        if raw_draw_number != draw_number:
                            logger.info(f"Basic normalization of draw number from '{raw_draw_number}' to '{draw_number}'")
                    
                    draw_date_str = row.get(column_mapping.get('draw_date', ''), None)
                    numbers_str = row.get(column_mapping.get('numbers', ''), None)
                    bonus_numbers_str = row.get(column_mapping.get('bonus_numbers', ''), None)
                    
                    # Skip rows with missing essential data
                    if not lottery_type or not draw_date_str or not numbers_str:
                        logger.warning(f"Skipping row {idx+2} due to missing key data")
                        continue
                    
                    # Parse date
                    draw_date = parse_date(draw_date_str)
                    if not draw_date:
                        logger.warning(f"Skipping row {idx+2} due to invalid date: {draw_date_str}")
                        continue
                    
                    # Parse numbers
                    numbers = parse_numbers(numbers_str)
                    if not numbers:
                        logger.warning(f"Skipping row {idx+2} due to invalid numbers: {numbers_str}")
                        continue
                    
                    # Ensure Daily Lottery has exactly 5 numbers
                    if lottery_type == "Daily Lottery":
                        if len(numbers) > 5:
                            logger.warning(f"Daily Lottery draw {draw_number} has {len(numbers)} numbers, limiting to first 5")
                            numbers = numbers[:5]
                        elif len(numbers) < 5:
                            logger.warning(f"Daily Lottery draw {draw_number} has only {len(numbers)} numbers, expected 5")
                    
                    # Parse bonus numbers
                    bonus_numbers = parse_numbers(bonus_numbers_str) if bonus_numbers_str else []
                    
                    # Handle divisions
                    divisions = {}
                    # First check if we have a dedicated divisions column
                    if 'divisions' in column_mapping and isinstance(column_mapping['divisions'], list):
                        # Check if we have division-specific columns (Div 1, Div 2, etc.)
                        div_columns = column_mapping['divisions']
                        
                        for div_col in div_columns:
                            div_value = row.get(div_col)
                            if not pd.isna(div_value):
                                # Try to extract division number from column name
                                div_name = div_col
                                if 'div' in str(div_col).lower():
                                    div_match = ''.join(c for c in str(div_col) if c.isdigit())
                                    if div_match:
                                        div_name = f"Division {div_match}"
                                
                                # If the value contains winners and prize together
                                if isinstance(div_value, str) and ',' in div_value:
                                    winners_part, prize_part = div_value.split(',', 1)
                                    # Extract just the number for winners
                                    winners = ''.join(c for c in winners_part if c.isdigit())
                                    # Keep the prize with formatting (R, commas)
                                    prize = prize_part.strip()
                                    divisions[div_name] = {
                                        "winners": winners,
                                        "prize": prize
                                    }
                                else:
                                    # If it's just a raw value, assume it's the prize amount
                                    divisions[div_name] = {
                                        "winners": "1",  # Default to 1 winner if not specified
                                        "prize": f"R{div_value}" if not str(div_value).startswith('R') else str(div_value)
                                    }
                    
                    # Check if this result already exists
                    existing = LotteryResult.query.filter_by(
                        lottery_type=lottery_type,
                        draw_number=draw_number
                    ).first()
                    
                    is_new_record = False
                    lottery_result = None

                    if existing:
                        # Update existing record
                        logger.info(f"Updating existing result for {lottery_type} draw {draw_number}")
                        existing.draw_date = draw_date
                        existing.numbers = json.dumps(numbers)
                        existing.bonus_numbers = json.dumps(bonus_numbers) if bonus_numbers else None
                        existing.divisions = json.dumps(divisions) if divisions else None
                        existing.source_url = "imported-from-excel"
                        existing.ocr_provider = "manual-import"
                        existing.ocr_model = "excel-spreadsheet"
                        existing.ocr_timestamp = datetime.utcnow().isoformat()
                        lottery_result = existing
                    else:
                        # Create new result
                        is_new_record = True
                        new_result = LotteryResult(
                            lottery_type=lottery_type,
                            draw_number=draw_number,
                            draw_date=draw_date,
                            numbers=json.dumps(numbers),
                            bonus_numbers=json.dumps(bonus_numbers) if bonus_numbers else None,
                            divisions=json.dumps(divisions) if divisions else None,
                            source_url="imported-from-excel",
                            ocr_provider="manual-import",
                            ocr_model="excel-spreadsheet",
                            ocr_timestamp=datetime.utcnow().isoformat()
                        )
                        db.session.add(new_result)
                        lottery_result = new_result
                    
                    # Commit each result individually to avoid losing all data if one fails
                    db.session.commit()
                    
                    # Store this record in import_tracking for later reference
                    # Import history will be created in main.py after this function returns
                    imported_records.append({
                        'lottery_type': lottery_type,
                        'draw_number': draw_number,
                        'draw_date': draw_date,
                        'is_new': is_new_record,
                        'lottery_result_id': lottery_result.id
                    })
                    
                    imported_count += 1
                    
                    if imported_count % 10 == 0:
                        logger.info(f"Imported {imported_count} records so far...")
                    
                except Exception as e:
                    errors_count += 1
                    logger.error(f"Error processing row {idx+2}: {str(e)}")
                    # Rollback this row only
                    db.session.rollback()
            
            # Count records after import
            final_count = LotteryResult.query.count()
            
            logger.info("Import completed!")
            logger.info(f"Records: {initial_count} -> {final_count} (added {final_count - initial_count})")
            logger.info(f"Successfully imported/updated {imported_count} records")
            logger.info(f"Encountered {errors_count} errors")
            
            # Return detailed import results
            return {
                'success': True,
                'total': imported_count,
                'initial_count': initial_count,
                'final_count': final_count,
                'added': final_count - initial_count,
                'errors': errors_count,
                'imported_records': imported_records  # Include records for import history
            }
    except Exception as e:
        logger.error(f"Error during import operation: {str(e)}")
        return False

def create_empty_template(output_path):
    """
    Create an empty Excel template for lottery data import.
    Single sheet format with standardized columns that matches the actual spreadsheet format.
    
    Args:
        output_path (str): Path to save the Excel file
    
    Returns:
        bool: Success status
    """
    try:
        # Define the columns based on observed format
        columns = [
            "Game Name", "Draw Number", "Draw Date", "Winning Numbers (Numerical)", "Bonus Ball",
            "Div 1 Winners", "Div 1 Winnings", "Div 2 Winners", "Div 2 Winnings",
            "Div 3 Winners", "Div 3 Winnings", "Div 4 Winners", "Div 4 Winnings",
            "Div 5 Winners", "Div 5 Winnings", "Div 6 Winners", "Div 6 Winnings",
            "Div 7 Winners", "Div 7 Winnings", "Div 8 Winners", "Div 8 Winnings",
            "Rollover Amount", "Total Pool Size", "Total Sales", "Next Jackpot",
            "Draw Machine", "Next Draw Date"
        ]
        
        # Create an empty dataframe with the columns
        df = pd.DataFrame(columns=columns)
        
        # Create Excel writer
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # Write to Excel without examples - completely empty
        df.to_excel(writer, sheet_name='Lotto', index=False)
        
        # Auto-adjust column widths
        workbook = writer.book
        worksheet = writer.sheets['Lotto']
        
        # Make columns wider for data entry
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(df.columns):
            col_width = max(len(col) + 5, 20)  # Minimum width of 20 characters
            col_letter = get_column_letter(i + 1)  # Convert number to Excel column letter (1=A, 27=AA, etc)
            worksheet.column_dimensions[col_letter].width = col_width
        
        # Add column formatting guidelines in first row
        # Get worksheet
        for i, col in enumerate(df.columns):
            cell = worksheet.cell(row=2, column=i+1)
            
            # Add guidance text based on column type
            if col == "Game Name":
                cell.value = "e.g., LOTTERY, POWERBALL, DAILY LOTTERY"
            elif col == "Draw Number":
                cell.value = "e.g., 2533"
            elif col == "Draw Date":
                cell.value = "YYYY-MM-DD format"
            elif col == "Winning Numbers (Numerical)":
                cell.value = "e.g., 01 02 03 04 05 06"
            elif col == "Bonus Ball":
                cell.value = "e.g., 07"
            elif "Winners" in col:
                cell.value = "Number of winners"
            elif "Winnings" in col:
                cell.value = "Prize amount (e.g., R1,000.00)"
        
        # Add light yellow background to guidance row
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        for i in range(len(df.columns)):
            worksheet.cell(row=2, column=i+1).fill = yellow_fill
        
        # Save the file
        writer.close()
        
        # Include timestamp in filename to avoid overwrites
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        print(f"Empty template created at {output_path}")
        logger.info(f"Empty template created at {output_path}")
        return True
    
    except Exception as e:
        logger.error(f"Error creating empty template: {str(e)}")
        print(f"Error creating empty template: {str(e)}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    if not os.path.exists(excel_file):
        print(f"Error: Excel file {excel_file} not found")
        sys.exit(1)
    
    # First purge existing data
    from purge_data import purge_data
    if purge_data():
        # Then import from Excel
        import_excel_data(excel_file)
    else:
        print("Error: Failed to purge existing data. Import aborted.")
        sys.exit(1)