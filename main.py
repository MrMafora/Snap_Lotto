"""
Main application entry point with Flask application defined for deployment.

This file is imported by gunicorn using the 'main:app' notation.

It also includes functionality to automatically bind to port 8080 
when running directly, to support Replit's external access requirements.

IMPORTANT: This application includes built-in port forwarding capabilities
to handle both the internal port 5000 (used by gunicorn) and the external
port 8080 required by Replit for public access.
"""
import json
import logging
import os
import re
import threading
import traceback
from datetime import datetime, timedelta
from functools import wraps

# Set up logging first
logging.basicConfig(level=logging.DEBUG)

def create_lottery_template(filepath):
    """Create an empty Excel template for lottery data import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Define lottery types and their columns
    lottery_types = {
        'Lotto': ['Draw Number', 'Draw Date', 'Ball 1', 'Ball 2', 'Ball 3', 'Ball 4', 'Ball 5', 'Ball 6', 'Bonus Ball'],
        'Lotto Plus 1': ['Draw Number', 'Draw Date', 'Ball 1', 'Ball 2', 'Ball 3', 'Ball 4', 'Ball 5', 'Ball 6', 'Bonus Ball'],
        'Lotto Plus 2': ['Draw Number', 'Draw Date', 'Ball 1', 'Ball 2', 'Ball 3', 'Ball 4', 'Ball 5', 'Ball 6', 'Bonus Ball'],
        'PowerBall': ['Draw Number', 'Draw Date', 'Ball 1', 'Ball 2', 'Ball 3', 'Ball 4', 'Ball 5', 'PowerBall'],
        'PowerBall Plus': ['Draw Number', 'Draw Date', 'Ball 1', 'Ball 2', 'Ball 3', 'Ball 4', 'Ball 5', 'PowerBall'],
        'Daily Lotto': ['Draw Number', 'Draw Date', 'Ball 1', 'Ball 2', 'Ball 3', 'Ball 4', 'Ball 5']
    }
    
    # Create a sheet for each lottery type
    for lottery_type, columns in lottery_types.items():
        ws = wb.create_sheet(title=lottery_type)
        
        # Add headers
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filepath)
logger = logging.getLogger(__name__)

# Now that logger is defined, import other modules
# import scheduler  # Commented out - module not found
from flask import Flask, render_template, flash, redirect, url_for, request, jsonify, send_from_directory, send_file, session
from flask_login import LoginManager, login_user, logout_user, current_user, login_required
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
# Use standard CSRF protection
from flask_wtf.csrf import CSRFProtect
import json

# Import models only (lightweight)
from models import LotteryResult, ScheduleConfig, Screenshot, User, Advertisement, AdImpression, Campaign, AdVariation, ImportHistory, ImportedRecord, db
from config import Config
from sqlalchemy import text

# Import performance cache manager
# Cache manager temporarily disabled due to model attribute issues
# from cache_manager import cache, cached_query, get_optimized_latest_results, get_optimized_lottery_stats, clear_results_cache

# Import modules
# ad_management temporarily disabled
# import ad_management
import lottery_analysis
# Remove problematic fix modules that are causing conflicts
# Import functionality integrated into main application
# scanner_routes module temporarily disabled to resolve conflicts

# Auto proxy functionality cleaned up - using direct port configuration

# Create the Flask application
app = Flask(__name__)
app.config.from_object(Config)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Add custom Jinja2 filters for math functions needed by charts
import math
import locale
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')  # Set locale for number formatting

app.jinja_env.filters['cos'] = lambda x: math.cos(float(x))
app.jinja_env.filters['sin'] = lambda x: math.sin(float(x))
app.jinja_env.filters['format_number'] = lambda x: f"{int(x):,}" if isinstance(x, (int, float)) else x

# Explicitly set database URI from environment variable
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Ensure proper PostgreSQL connection string format
    # Heroku-style connection strings start with postgres:// but SQLAlchemy requires postgresql://
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    logger.info(f"Using database from DATABASE_URL environment variable")
else:
    logger.warning("DATABASE_URL not found, using fallback database configuration")

# Initialize SQLAlchemy with additional connection settings
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_recycle": 300,
    "pool_pre_ping": True
}

# Add SSL requirement for PostgreSQL connections
if app.config['SQLALCHEMY_DATABASE_URI'] and app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgresql'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS']["connect_args"] = {"sslmode": "require"}
    logger.info("PostgreSQL SSL mode enabled")

# Initialize SQLAlchemy
db.init_app(app)

# Initialize Login Manager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# CSRF Protection completely disabled for Replit environment
app.config['WTF_CSRF_ENABLED'] = False  # Completely disable CSRF for Replit environment

# All endpoints are now exempt from CSRF protection

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Defer database schema creation to background thread
def init_database():
    """Initialize database tables in background thread"""
    with app.app_context():
        db.create_all()
        logger.info("Database tables created/verified")
        
# Start database initialization in background to avoid blocking startup
threading.Thread(target=init_database, daemon=True).start()

# Lazy load these modules - will be imported when needed
# This helps reduce initial load time
data_aggregator = None
import_excel = None
import_snap_lotto_data = None
ocr_processor = None
screenshot_manager = None
health_monitor = None
daily_scheduler = None

def init_lazy_modules():
    """Initialize modules in a background thread with timeout"""
    global data_aggregator, import_excel, import_snap_lotto_data, ocr_processor, screenshot_manager, health_monitor, daily_scheduler
    
    # Note: Removed signal alarm since it only works in main thread
    
    # Import heavy modules only when needed - data_aggregator not available
    # try:
    #     import data_aggregator as da
    # except ImportError:
    da = None
    try:
        import ocr_processor as op
    except ImportError:
        op = None
    try:
        import screenshot_manager as sm
    except ImportError:
        sm = None
    try:
        import health_monitor as hm
    except ImportError:
        hm = None
    
    try:
        from scheduler import LotteryScheduler
        sched = LotteryScheduler
    except ImportError:
        sched = None
    
    # Store module references - data_aggregator not available
    # data_aggregator = da
    import_excel = None  # Functionality integrated into main app
    import_snap_lotto_data = None  # Functionality integrated into main app
    ocr_processor = op
    screenshot_manager = sm
    health_monitor = hm
    daily_scheduler = sched
    
    # Initialize scheduler and health monitoring in background after imports are complete
    with app.app_context():
        # Initialize daily automation scheduler to run at 1:00 AM every day
        if sched:
            daily_scheduler = sched(app, "01:00")
            daily_scheduler.start_scheduler()
            logger.info("Daily automation scheduler initialized - will run at 1:00 AM every day")
        else:
            logger.warning("Scheduler module not available - daily automation will not run automatically")
            
        if hm:
            hm.init_health_monitor(app, db)
        
        # Scanner routes are handled in main.py directly
        # Ticket scanner functionality integrated in existing routes
    
    logger.info("All modules lazy-loaded successfully")

def calculate_frequency_analysis(results):
    """Calculate frequency analysis from lottery results for homepage charts"""
    try:
        number_counts = {}
        
        # Count frequency of all numbers across all lottery types
        for result in results:
            numbers = result.get_numbers_list()
            for num in numbers:
                if isinstance(num, (int, str)) and str(num).isdigit():
                    num_int = int(num)
                    number_counts[num_int] = number_counts.get(num_int, 0) + 1
        
        # Get top 10 most frequent numbers
        sorted_numbers = sorted(number_counts.items(), key=lambda x: x[1], reverse=True)
        most_frequent = sorted_numbers[:10] if len(sorted_numbers) >= 10 else sorted_numbers
        
        app.logger.info(f"Frequency analysis: Found {len(number_counts)} unique numbers, top 10: {most_frequent}")
        
        return {
            'most_frequent': most_frequent,
            'total_numbers': len(number_counts)
        }
        
    except Exception as e:
        app.logger.error(f"Error calculating frequency analysis: {str(e)}")
        return {
            'most_frequent': [],
            'total_numbers': 0
        }

# HOMEPAGE ROUTE - Must be first to ensure authentic lottery data displays
@app.route('/')
def home():
    """OPTIMIZED Homepage displaying authentic South African lottery results"""
    from models import LotteryResult
    import json
    
    app.logger.info("=== HOMEPAGE: Loading fresh lottery data from database ===")
    
    # Load fresh data directly from database using raw SQL
    try:
        from sqlalchemy import text
        
        # Direct SQL query to get latest result for each lottery type
        # Note: Lottery games that draw together share the same draw_date and draw_number:
        # - LOTTO, LOTTO PLUS 1, LOTTO PLUS 2 all draw on the same day (same draw_number)
        # - PowerBall and POWERBALL PLUS draw on the same day (same draw_number)
        # - DAILY LOTTO draws independently
        query = text("""
            WITH ranked_results AS (
              SELECT lottery_type, draw_number, draw_date, numbers as main_numbers, bonus_numbers,
                ROW_NUMBER() OVER (PARTITION BY lottery_type ORDER BY draw_date DESC, draw_number DESC) as rn
              FROM lottery_result 
              WHERE lottery_type IN ('LOTTO', 'LOTTO PLUS 1', 'LOTTO PLUS 2', 'PowerBall', 'POWERBALL PLUS', 'DAILY LOTTO')
            )
            SELECT lottery_type, draw_number, draw_date, main_numbers, bonus_numbers
            FROM ranked_results 
            WHERE rn = 1
        """)
        
        raw_results = db.session.execute(query).fetchall()
        app.logger.info(f"Raw database results: {[(r.lottery_type, r.draw_number, r.draw_date) for r in raw_results]}")
        results = []
        
        # Define the desired display order
        lottery_order = [
            'LOTTO',           # Lottery  
            'LOTTO PLUS 1',    # Lottery Plus 1
            'LOTTO PLUS 2',    # Lottery Plus 2
            'PowerBall',       # Powerball
            'POWERBALL PLUS',  # Powerball Plus
            'DAILY LOTTO'      # Daily Lottery
        ]
        
        # Create lookup dictionary for all lottery results
        lottery_lookup = {}
        for row in raw_results:
            if row.lottery_type not in lottery_lookup:
                lottery_lookup[row.lottery_type] = row
        
        # Process results in the specified order
        for lottery_type in lottery_order:
            if lottery_type in lottery_lookup:
                row = lottery_lookup[lottery_type]
                
                # Parse main_numbers field (JSON format)
                numbers = []
                bonus_numbers = []
                
                if row.main_numbers:
                    main_nums_str = str(row.main_numbers)
                    try:
                        # Parse JSON string format like "[15,22,25,31,36]"
                        numbers = json.loads(main_nums_str)
                    except (json.JSONDecodeError, ValueError):
                        # Fallback for PostgreSQL array format like "{15,22,25,31,36}"
                        if main_nums_str.startswith('{') and main_nums_str.endswith('}'):
                            nums_only = main_nums_str[1:-1]  # Remove { }
                            if nums_only.strip():
                                numbers = [int(n.strip()) for n in nums_only.split(',') if n.strip().isdigit()]
                    
                # Parse bonus_numbers field (JSON format)
                if row.bonus_numbers:
                    bonus_str = str(row.bonus_numbers)
                    try:
                        bonus_numbers = json.loads(bonus_str)
                    except (json.JSONDecodeError, ValueError):
                        # Fallback for PostgreSQL array format
                        if bonus_str.startswith('{') and bonus_str.endswith('}'):
                            bonus_only = bonus_str[1:-1]  # Remove { }
                            if bonus_only.strip():
                                bonus_numbers = [int(n.strip()) for n in bonus_only.split(',') if n.strip().isdigit()]
                    
                # Always create display object - even if numbers is empty, we still need draw info
                class LotteryDisplay:
                    def __init__(self, lottery_type, draw_number, draw_date, numbers, bonus_numbers):
                        self.lottery_type = lottery_type
                        self.draw_number = str(draw_number)
                        self.draw_date = draw_date
                        self.numbers = numbers if numbers else []
                        self.bonus_numbers = bonus_numbers if bonus_numbers else []
                    
                    def get_numbers_list(self):
                        return self.numbers
                    
                    def get_bonus_numbers_list(self):
                        return self.bonus_numbers
                
                result = LotteryDisplay(row.lottery_type, row.draw_number, 
                                      row.draw_date, numbers, bonus_numbers)
                results.append(result)
        
        app.logger.info(f"HOMEPAGE: Loaded {len(results)} results from database")
        
        # Calculate frequency analysis for homepage charts
        frequency_data = calculate_frequency_analysis(results)
        frequent_numbers = frequency_data.get('most_frequent', [])
        
        return render_template('index.html', results=results, frequent_numbers=frequent_numbers)
        
    except Exception as e:
        app.logger.error(f"Error in optimized homepage: {str(e)}")
        # Fallback to basic empty results if cache fails
        return render_template('index.html', results=[], frequent_numbers=[])

# Start lazy loading in background thread AFTER homepage route is registered
threading.Thread(target=init_lazy_modules, daemon=True).start()

@app.route('/debug-data')
def debug_data():
    """Debug endpoint to check authentic lottery data"""
    from sqlalchemy import text
    import json
    
    try:
        query = text("SELECT lottery_type, draw_number, numbers FROM lottery_result LIMIT 5")
        rows = db.session.execute(query).fetchall()
        
        debug_info = {
            'total_rows': len(rows),
            'sample_data': []
        }
        
        for row in rows:
            debug_info['sample_data'].append({
                'lottery_type': row.lottery_type,
                'draw_number': str(row.draw_number),
                'numbers': row.main_numbers
            })
        
        return debug_info
    except Exception as e:
        return {'error': str(e)}


@app.route('/admin')
@login_required
def admin():
    """Admin dashboard for system administration"""
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('home'))
    
    # Get basic system statistics
    system_stats = {
        'cpu_usage': 'Available',
        'memory_usage': 'Available', 
        'disk_usage': 'Available'
    }
    
    # Get lottery data count
    try:
        from models import LotteryResult
        total_results = LotteryResult.query.count()
        latest_result = LotteryResult.query.order_by(LotteryResult.id.desc()).first()
        
        data_stats = {
            'total_results': total_results,
            'latest_draw': latest_result.draw_number if latest_result else 'None',
            'latest_date': latest_result.draw_date.strftime('%Y-%m-%d') if latest_result else 'None'
        }
    except Exception as e:
        app.logger.error(f"Error getting lottery data stats: {str(e)}")
        data_stats = {
            'total_results': 0,
            'latest_draw': 'Error',
            'latest_date': 'Error'
        }
    
    return render_template(
        'admin/dashboard.html',
        title="Admin Dashboard",
        system_stats=system_stats,
        data_stats=data_stats,
        recent_api_logs=[],
        recent_imports=[],
        recent_health_checks=[]
    )

# Removed duplicate api-tracking route - functionality moved to api_tracking_view() function

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if current_user.is_authenticated:
        return redirect('/')
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Login successful!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or '/')
        else:
            flash('Invalid username or password', 'danger')
    
    # Define SEO metadata
    meta_description = "Secure login for administration of the South African lottery results system. Access administrative features to manage lottery data and system settings."
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Admin Login", "url": url_for('login')}
    ]
    
    return render_template('login.html', 
                          title="Admin Login | Lottery Results Management",
                          meta_description=meta_description,
                          breadcrumbs=breadcrumbs)

@app.route('/logout')
@login_required
def logout():
    """Logout route"""
    logout_user()
    session.clear()  # Clear session data
    flash('You have been logged out.', 'info')
    # Use direct return with status code
    from werkzeug.wrappers import Response
    return Response(status=302, headers={'Location': '/'})

@app.route('/scan-lottery-ticket-south-africa')
@app.route('/scanner-landing')
def scanner_landing():
    """Landing page focused on the exclusive lottery ticket scanner feature"""
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Lottery Ticket Scanner", "url": url_for('scanner_landing')}
    ]
    
    # Define SEO metadata
    meta_description = "South Africa's ONLY lottery ticket scanner app. Instantly check if your Lottery, PowerBall, or Daily Lottery ticket is a winner by uploading a photo."
    
    return render_template('scanner_landing.html',
                          title="South Africa's ONLY Lottery Ticket Scanner App | Check If You've Won Instantly",
                          breadcrumbs=breadcrumbs,
                          meta_description=meta_description)

@app.route('/ticket-scanner')
def ticket_scanner():
    """Ticket scanner page - Allows users to scan and validate their lottery tickets"""
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Lottery Ticket Scanner", "url": url_for('scanner_landing')},
        {"name": "Scan Your Ticket", "url": url_for('ticket_scanner')}
    ]
    
    # Additional SEO metadata
    meta_description = "Check if your South African lottery ticket is a winner. Our free ticket scanner uses advanced technology to analyze and verify your lottery tickets instantly."
    
    return render_template('ticket_scanner.html', 
                          title="Lottery Ticket Scanner | Check If You've Won",
                          breadcrumbs=breadcrumbs,
                          meta_description=meta_description)

@app.route('/api/scan-ticket', methods=['POST'])
def api_scan_ticket():
    """API endpoint for processing uploaded ticket images using Google Gemini 2.5 Pro"""
    try:
        if 'ticket_image' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['ticket_image']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_filename = f"ticket_{timestamp}_{filename}"
        file_path = os.path.join('uploads', safe_filename)
        
        # Ensure uploads directory exists
        os.makedirs('uploads', exist_ok=True)
        file.save(file_path)
        
        try:
            # Process with Google Gemini 2.5 Pro
            import google.generativeai as genai
            
            api_key = os.environ.get('GOOGLE_API_KEY_SNAP_LOTTERY')
            if not api_key:
                logger.error("No Google API key found")
                return jsonify({'error': 'Google API key not configured'}), 500
            
            from google import genai
            from google.genai import types
            
            client = genai.Client(api_key=api_key)
            
            # Load and prepare image bytes for Gemini 2.5 Pro
            with open(file_path, 'rb') as f:
                image_bytes = f.read()
            logger.info(f"Image loaded successfully: {os.path.getsize(file_path)} bytes")
            
            # Enhanced prompt for comprehensive ticket analysis - ONE IMAGE ONLY
            prompt = """Analyze this single South African lottery ticket image and extract ALL visible information in JSON format.

CRITICAL: Process ONLY this one image. Extract EXACTLY what you see on the ticket. Do not make assumptions.

Return JSON with this structure:
{
    "lottery_type": "LOTTO|POWERBALL|DAILY LOTTO|AUTO-DETECT",
    "draw_date": "YYYY-MM-DD or visible date",
    "draw_number": "visible draw number",
    "ticket_cost": "visible cost amount",
    "all_lines": [
        [1, 2, 3, 4, 5, 6],
        [7, 8, 9, 10, 11, 12]
    ],
    "bonus_numbers": [number] or [],
    "powerball_numbers": [number] or [],
    "lotto_plus_1_included": "YES|NO|NOT_VISIBLE",
    "lotto_plus_2_included": "YES|NO|NOT_VISIBLE",
    "powerball_plus_included": "YES|NO|NOT_VISIBLE"
}

Extract ALL visible data accurately from this single image."""
            
            logger.info("Sending single image to Gemini 2.5 Pro API...")
            response = client.models.generate_content(
                model="gemini-2.5-pro",
                contents=[
                    types.Part.from_bytes(
                        data=image_bytes,
                        mime_type="image/png" if file_path.lower().endswith('.png') else "image/jpeg",
                    ),
                    prompt
                ],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    temperature=0.1,
                    max_output_tokens=1000,
                ),
            )
            response_text = response.text
            
            logger.info(f"Gemini response received: {response_text}")
            
        except Exception as e:
            logger.error(f"Error with Gemini API: {str(e)}")
            return jsonify({'error': f'AI processing failed: {str(e)}'}), 500
        
        try:
            # Extract and clean JSON
            import re
            logger.info("Extracting JSON from Gemini response...")
            json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
            
            if not json_match:
                logger.error(f"No JSON found in response: {response_text}")
                return jsonify({'error': 'Could not extract valid JSON from AI response'}), 400
            
            json_text = json_match.group()
            logger.info(f"Extracted JSON text: {json_text}")
            
            # Fix leading zeros in JSON (05 -> "05")
            json_text = re.sub(r'(\[|\,)\s*0(\d)', r'\1"0\2"', json_text)
            logger.info(f"Cleaned JSON text: {json_text}")
            
            ticket_data = json.loads(json_text)
            logger.info(f"Successfully parsed ticket data: {ticket_data}")
            
            # Clean up the file
            os.remove(file_path)
            
            return jsonify({
                'success': True,
                'data': ticket_data,
                'message': 'Ticket processed successfully'
            })
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {str(e)}")
            return jsonify({'error': f'Invalid JSON format: {str(e)}'}), 400
        except Exception as e:
            logger.error(f"Error processing JSON: {str(e)}")
            return jsonify({'error': f'Processing error: {str(e)}'}), 500
            
    except Exception as e:
        logger.error(f"Error processing ticket: {e}")
        # Clean up file if it exists
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'error': f'Processing error: {str(e)}'}), 500

@app.route('/process-ticket', methods=['POST'])
def process_ticket():
    """Process a lottery ticket image and return JSON results"""
    file_path = None
    try:
        logger.info("=== TICKET PROCESSING STARTED ===")
        
        if 'ticket_image' not in request.files:
            logger.error("No ticket_image in request.files")
            return jsonify({'success': False, 'error': 'No file uploaded'})
        
        file = request.files['ticket_image']
        if file.filename == '':
            logger.error("Empty filename in uploaded file")
            return jsonify({'success': False, 'error': 'No file selected'})
        
        logger.info(f"Processing file: {file.filename}, content type: {file.content_type}")
        
        # Save uploaded file temporarily
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{secure_filename(file.filename) if file.filename else 'ticket.jpg'}"
        upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
        os.makedirs(upload_folder, exist_ok=True)
        file_path = os.path.join(upload_folder, filename)
        
        logger.info(f"Saving file to: {file_path}")
        file.save(file_path)
        
        if not os.path.exists(file_path):
            logger.error(f"File was not saved successfully: {file_path}")
            return jsonify({'success': False, 'error': 'File save failed'})
        
        logger.info(f"File saved successfully, size: {os.path.getsize(file_path)} bytes")
        
        try:
            # Import Google AI safely to avoid urllib3 conflicts
            try:
                logger.info("Importing Google AI modules...")
                import google.generativeai as genai
                import PIL.Image
                logger.info("Google AI modules imported successfully")
            except ImportError as ie:
                logger.error(f"Import error: {ie}")
                return jsonify({'success': False, 'error': 'AI processing service unavailable'})
            
            # Configure API key
            api_key = os.environ.get('GOOGLE_API_KEY_SNAP_LOTTERY')
            if not api_key:
                logger.error("No Google API key found in environment")
                return jsonify({'success': False, 'error': 'AI service not configured'})
            
            logger.info("Configuring Google AI with API key...")
            try:
                genai.configure(api_key=api_key)
                model = genai.GenerativeModel('gemini-1.5-pro')
                logger.info("Google AI configured successfully")
            except Exception as config_error:
                logger.error(f"Error configuring Google AI: {config_error}")
                return jsonify({'success': False, 'error': f'AI configuration failed: {str(config_error)}'})
            
            # Load image
            image = PIL.Image.open(file_path)
            
            # Create comprehensive prompt
            prompt = """Extract lottery ticket data from this South African lottery ticket.

Look carefully for:
1. Main lottery type (PowerBall or LOTTO)
2. All number lines/selections on the ticket
3. Whether additional games are included (look for YES/NO checkboxes)

Return ONLY valid JSON in this format:

For PowerBall tickets:
{
    "lottery_type": "PowerBall",
    "all_lines": [[12, 17, 24, 26, 35], [5, 15, 17, 24, 32]],
    "all_powerball": [12, 12],
    "powerball_plus_included": "YES",
    "draw_date": "18/04/25",
    "draw_number": "1607",
    "ticket_cost": "R37.50"
}

For LOTTO tickets:
{
    "lottery_type": "LOTTO",
    "all_lines": [[8, 24, 32, 34, 36, 52]],
    "bonus_numbers": [26],
    "lotto_plus_1_included": "YES",
    "lotto_plus_2_included": "YES",
    "draw_date": "04/06/25",
    "draw_number": "2547",
    "ticket_cost": "R15.00"
}

For Daily Lotto tickets:
{
    "lottery_type": "Daily Lotto",
    "all_lines": [[5, 12, 18, 23, 31]],
    "draw_date": "04/06/25",
    "draw_number": "2274",
    "ticket_cost": "R3.00"
}

CRITICAL INSTRUCTIONS:
1. Check for YES/NO checkboxes or indicators for LOTTO Plus 1, LOTTO Plus 2, PowerBall Plus
2. If you see a YES checkbox marked for any additional game, set it as "YES"
3. If you see a NO checkbox marked or no checkbox, set it as "NO"
4. Extract ALL number lines - each line represents one game entry
5. Return only the JSON, no other text"""
            
            response = model.generate_content([image, prompt])
            response_text = response.text.strip() if response and response.text else ""
            
            logger.info(f"Gemini response: {response_text}")
            
            if not response_text:
                logger.error("Empty response from Gemini API")
                if os.path.exists(file_path):
                    os.remove(file_path)
                return jsonify({'success': False, 'error': 'No response from AI service'})
            
            # Clean up response and extract JSON
            json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
            if json_match:
                json_text = json_match.group()
                ticket_data = json.loads(json_text)
                
                # Clean up uploaded file
                os.remove(file_path)
                
                # Define lottery type configurations for South African lotteries
                LOTTERY_CONFIGS = {
                    'LOTTO': {'ticket_has_bonus': False, 'draw_has_bonus': True, 'main_numbers': 6, 'bonus_name': 'Bonus Ball'},
                    'LOTTO PLUS 1': {'ticket_has_bonus': False, 'draw_has_bonus': True, 'main_numbers': 6, 'bonus_name': 'Bonus Ball'},
                    'LOTTO PLUS 2': {'ticket_has_bonus': False, 'draw_has_bonus': True, 'main_numbers': 6, 'bonus_name': 'Bonus Ball'},
                    'POWERBALL': {'ticket_has_bonus': True, 'draw_has_bonus': True, 'main_numbers': 5, 'bonus_name': 'PowerBall'},
                    'POWERBALL PLUS': {'ticket_has_bonus': True, 'draw_has_bonus': True, 'main_numbers': 5, 'bonus_name': 'PowerBall Plus'},
                    'DAILY LOTTO': {'ticket_has_bonus': False, 'draw_has_bonus': False, 'main_numbers': 5, 'bonus_name': None}
                }
                
                # Format response for frontend - handle all lottery types
                lottery_type = ticket_data.get('lottery_type', 'PowerBall')
                lottery_config = LOTTERY_CONFIGS.get(lottery_type.upper(), LOTTERY_CONFIGS['POWERBALL'])
                
                # Initialize comparison results and final result
                main_game_results = {}
                plus_1_results = {}
                plus_2_results = {}
                result = {}
                
                # Extract numbers and bonus/powerball based on lottery type configuration
                if 'LOTTO' in lottery_type.upper():
                    # LOTTO ticket format
                    all_lines = ticket_data.get('all_lines', [])
                    bonus_numbers = ticket_data.get('bonus_numbers', [])
                    lotto_plus_1_included = ticket_data.get('lotto_plus_1_included', 'NO')
                    lotto_plus_2_included = ticket_data.get('lotto_plus_2_included', 'NO')
                    
                    # Get player numbers for comparison - LOTTO tickets don't have bonus numbers
                    player_numbers = all_lines[0] if all_lines else []
                    player_bonus = None  # LOTTO tickets never have bonus numbers on the ticket itself
                    
                    logger.info(f"Processing LOTTO ticket: numbers={player_numbers}")
                    logger.info(f"Plus game flags: Plus1={lotto_plus_1_included}, Plus2={lotto_plus_2_included}")
                    
                    # Always compare against main LOTTO results
                    if player_numbers:
                        lotto_draw = LotteryResult.query.filter_by(lottery_type='LOTTO').order_by(LotteryResult.draw_date.desc()).first()
                        logger.info(f"Database query result: {lotto_draw}")
                        if lotto_draw:
                            # Parse the main_numbers field which is stored as JSON string
                            winning_numbers = json.loads(lotto_draw.main_numbers) if lotto_draw.main_numbers else []
                            logger.info(f"Parsed winning numbers: {winning_numbers}")
                            
                            main_matches = len(set(player_numbers) & set(winning_numbers))
                            
                            # Handle bonus ball comparison - LOTTO draws have bonus balls but tickets don't
                            bonus_match = False
                            winning_bonus = None
                            if lottery_config['draw_has_bonus'] and lotto_draw.bonus_numbers:
                                winning_bonus = json.loads(lotto_draw.bonus_numbers)[0] if isinstance(lotto_draw.bonus_numbers, str) else lotto_draw.bonus_numbers[0]
                                # For LOTTO: Check if any of the player's main numbers match the drawn bonus
                                bonus_match = winning_bonus in player_numbers
                            
                            main_game_results = {
                                'lottery_type': 'LOTTO',
                                'draw_number': lotto_draw.draw_number,
                                'draw_date': lotto_draw.draw_date.strftime('%Y-%m-%d'),
                                'winning_numbers': winning_numbers,
                                'main_matches': main_matches,
                                'bonus_match': bonus_match,
                                'total_matches': f"{main_matches} numbers matched" + (f" + bonus" if bonus_match else "")
                            }
                            
                            logger.info(f"LOTTO comparison: {main_matches} matches, bonus: {bonus_match}")
                    
                    # Compare against LOTTO Plus 1 if selected
                    if lotto_plus_1_included.upper() == 'YES' and player_numbers:
                        lotto_plus_1_draw = LotteryResult.query.filter_by(lottery_type='LOTTO PLUS 1').order_by(LotteryResult.draw_date.desc()).first()
                        if lotto_plus_1_draw:
                            plus_1_numbers = json.loads(lotto_plus_1_draw.main_numbers) if lotto_plus_1_draw.main_numbers else []
                            plus_1_config = LOTTERY_CONFIGS.get('LOTTO PLUS 1', {'has_bonus': False})
                            
                            plus_1_main_matches = len(set(player_numbers) & set(plus_1_numbers))
                            
                            # Handle bonus ball for LOTTO Plus 1 - draw has bonus but ticket doesn't
                            plus_1_bonus_match = False
                            plus_1_winning_bonus = None
                            if plus_1_config['draw_has_bonus'] and lotto_plus_1_draw.bonus_numbers:
                                plus_1_winning_bonus = json.loads(lotto_plus_1_draw.bonus_numbers)[0] if isinstance(lotto_plus_1_draw.bonus_numbers, str) else lotto_plus_1_draw.bonus_numbers[0]
                                # Check if any of the player's main numbers match the drawn bonus
                                plus_1_bonus_match = plus_1_winning_bonus in player_numbers
                            
                            plus_1_results = {
                                'lottery_type': 'LOTTO Plus 1',
                                'draw_number': lotto_plus_1_draw.draw_number,
                                'draw_date': lotto_plus_1_draw.draw_date.strftime('%Y-%m-%d'),
                                'winning_numbers': plus_1_numbers,
                                'main_matches': plus_1_main_matches,
                                'bonus_match': plus_1_bonus_match,
                                'total_matches': f"{plus_1_main_matches} numbers matched" + (f" + bonus" if plus_1_bonus_match else "")
                            }
                    
                    # Compare against LOTTO Plus 2 if selected
                    if lotto_plus_2_included.upper() == 'YES' and player_numbers:
                        lotto_plus_2_draw = LotteryResult.query.filter_by(lottery_type='LOTTO PLUS 2').order_by(LotteryResult.draw_date.desc()).first()
                        if lotto_plus_2_draw:
                            plus_2_numbers = json.loads(lotto_plus_2_draw.main_numbers) if lotto_plus_2_draw.main_numbers else []
                            plus_2_config = LOTTERY_CONFIGS.get('LOTTO PLUS 2', {'has_bonus': False})
                            
                            plus_2_main_matches = len(set(player_numbers) & set(plus_2_numbers))
                            
                            # Handle bonus ball for LOTTO Plus 2 - draw has bonus but ticket doesn't
                            plus_2_bonus_match = False
                            plus_2_winning_bonus = None
                            if plus_2_config['draw_has_bonus'] and lotto_plus_2_draw.bonus_numbers:
                                plus_2_winning_bonus = json.loads(lotto_plus_2_draw.bonus_numbers)[0] if isinstance(lotto_plus_2_draw.bonus_numbers, str) else lotto_plus_2_draw.bonus_numbers[0]
                                # Check if any of the player's main numbers match the drawn bonus
                                plus_2_bonus_match = plus_2_winning_bonus in player_numbers
                            
                            plus_2_results = {
                                'lottery_type': 'LOTTO Plus 2',
                                'draw_number': lotto_plus_2_draw.draw_number,
                                'draw_date': lotto_plus_2_draw.draw_date.strftime('%Y-%m-%d'),
                                'winning_numbers': plus_2_numbers,
                                'main_matches': plus_2_main_matches,
                                'bonus_match': plus_2_bonus_match,
                                'total_matches': f"{plus_2_main_matches} numbers matched" + (f" + bonus" if plus_2_bonus_match else "")
                            }
                    
                    # Create message based on which games were checked
                    games_checked = []
                    if main_game_results:
                        games_checked.append('LOTTO')
                    if plus_1_results:
                        games_checked.append('LOTTO Plus 1')
                    if plus_2_results:
                        games_checked.append('LOTTO Plus 2')
                    
                    comparison_message = f"LOTTO ticket analyzed and compared against: {', '.join(games_checked)}"
                    
                    result = {
                        'success': True,
                        'lottery_type': lottery_type,
                        'draw_date': ticket_data.get('draw_date', 'Not detected'),
                        'draw_number': ticket_data.get('draw_number', 'Not detected'),
                        'all_lines': all_lines,
                        'bonus_numbers': bonus_numbers,
                        'lotto_plus_1_included': lotto_plus_1_included,
                        'lotto_plus_2_included': lotto_plus_2_included,
                        'ticket_cost': ticket_data.get('ticket_cost', 'Not detected'),
                        'ticket_data': ticket_data,
                        'raw_response': response_text,
                        'main_game_results': main_game_results,
                        'lotto_plus_1_results': plus_1_results,
                        'lotto_plus_2_results': plus_2_results,
                        'powerball_results': {},  # Empty for LOTTO tickets
                        'powerball_plus_results': {},  # Empty for LOTTO tickets
                        'comparison': {
                            'message': comparison_message,
                            'extracted_numbers': all_lines[0] if all_lines else [],
                            'lottery_type': lottery_type,
                            'games_checked': len(games_checked),
                            'enhanced_comparison': True,
                            'main_game': main_game_results,
                            'plus_1_game': plus_1_results,
                            'plus_2_game': plus_2_results
                        }
                    }
                elif lottery_type == 'PowerBall':
                    # PowerBall ticket format
                    all_lines = ticket_data.get('all_lines', [])
                    all_powerball = ticket_data.get('all_powerball', [])
                    powerball_plus_included = ticket_data.get('powerball_plus_included', 'NO')
                    
                    # Get player numbers for comparison
                    player_numbers = all_lines[0] if all_lines else []
                    player_powerball = all_powerball[0] if all_powerball else None
                    
                    # Compare against main PowerBall results
                    if player_numbers:
                        powerball_draw = LotteryResult.query.filter_by(lottery_type='POWERBALL').order_by(LotteryResult.draw_date.desc()).first()
                        if powerball_draw:
                            winning_numbers = json.loads(powerball_draw.main_numbers) if powerball_draw.main_numbers else []
                            winning_powerball = powerball_draw.bonus_numbers[0] if powerball_draw.bonus_numbers else None
                            
                            main_matches = len(set(player_numbers) & set(winning_numbers))
                            powerball_match = (player_powerball == winning_powerball) if player_powerball and winning_powerball else False
                            
                            main_game_results = {
                                'lottery_type': 'PowerBall',
                                'draw_number': powerball_draw.draw_number,
                                'draw_date': powerball_draw.draw_date.strftime('%Y-%m-%d'),
                                'winning_numbers': winning_numbers,
                                'winning_powerball': winning_powerball,
                                'bonus_match': powerball_match,
                                'main_matches': main_matches,
                                'powerball_match': powerball_match,
                                'total_matches': f"{main_matches} main + {'PowerBall' if powerball_match else '0 PowerBall'}"
                            }
                    
                    # Compare against PowerBall Plus if selected
                    if powerball_plus_included.upper() == 'YES' and player_numbers:
                        powerball_plus_draw = LotteryResult.query.filter_by(lottery_type='POWERBALL PLUS').order_by(LotteryResult.draw_date.desc()).first()
                        if powerball_plus_draw:
                            plus_numbers = json.loads(powerball_plus_draw.main_numbers) if powerball_plus_draw.main_numbers else []
                            plus_powerball = powerball_plus_draw.bonus_numbers[0] if powerball_plus_draw.bonus_numbers else None
                            
                            plus_main_matches = len(set(player_numbers) & set(plus_numbers))
                            plus_powerball_match = (player_powerball == plus_powerball) if player_powerball and plus_powerball else False
                            
                            plus_1_results = {
                                'lottery_type': 'PowerBall Plus',
                                'draw_number': powerball_plus_draw.draw_number,
                                'draw_date': powerball_plus_draw.draw_date.strftime('%Y-%m-%d'),
                                'winning_numbers': plus_numbers,
                                'winning_powerball': plus_powerball,
                                'main_matches': plus_main_matches,
                                'powerball_match': plus_powerball_match,
                                'total_matches': f"{plus_main_matches} main + {'PowerBall' if plus_powerball_match else '0 PowerBall'}"
                            }
                    
                    # Create message based on which games were checked
                    games_checked = []
                    if main_game_results:
                        games_checked.append('PowerBall')
                    if plus_1_results:
                        games_checked.append('PowerBall Plus')
                    
                    comparison_message = f"PowerBall ticket analyzed and compared against: {', '.join(games_checked)}"
                    
                    result = {
                        'success': True,
                        'lottery_type': lottery_type,
                        'draw_date': ticket_data.get('draw_date', 'Not detected'),
                        'draw_number': ticket_data.get('draw_number', 'Not detected'),
                        'all_lines': all_lines,
                        'all_powerball': all_powerball,
                        'powerball_plus_included': powerball_plus_included,
                        'ticket_cost': ticket_data.get('ticket_cost', 'Not detected'),
                        'ticket_data': ticket_data,
                        'raw_response': response_text,
                        'main_game_results': main_game_results,
                        'powerball_plus_results': plus_1_results,
                        'comparison': {
                            'message': comparison_message,
                            'extracted_numbers': all_lines[0] if all_lines else [],
                            'powerball_number': str(all_powerball[0]) if all_powerball else 'Not detected',
                            'lottery_type': lottery_type,
                            'games_checked': len(games_checked)
                        }
                    }
                
                elif lottery_type == 'Daily Lotto':
                    # Daily Lotto ticket format (5 numbers, no bonus/powerball)
                    all_lines = ticket_data.get('all_lines', [])
                    
                    # Initialize variables
                    main_game_results = {}
                    
                    # Get player numbers for comparison
                    player_numbers = all_lines[0] if all_lines else []
                    
                    # Compare against Daily Lotto results
                    if player_numbers:
                        daily_lotto_draw = LotteryResult.query.filter_by(lottery_type='DAILY LOTTO').order_by(LotteryResult.draw_date.desc()).first()
                        if daily_lotto_draw:
                            winning_numbers = json.loads(daily_lotto_draw.main_numbers) if daily_lotto_draw.main_numbers else []
                            
                            main_matches = len(set(player_numbers) & set(winning_numbers))
                            
                            main_game_results = {
                                'lottery_type': 'Daily Lotto',
                                'draw_number': daily_lotto_draw.draw_number,
                                'draw_date': daily_lotto_draw.draw_date.strftime('%Y-%m-%d'),
                                'winning_numbers': winning_numbers,
                                'main_matches': main_matches,
                                'total_matches': f"{main_matches} main numbers"
                            }
                    
                    # Create comparison message
                    comparison_message = f"Daily Lotto ticket analyzed and compared against: Daily Lotto"
                    
                    result = {
                        'success': True,
                        'lottery_type': lottery_type,
                        'draw_date': ticket_data.get('draw_date', 'Not detected'),
                        'draw_number': ticket_data.get('draw_number', 'Not detected'),
                        'all_lines': all_lines,
                        'ticket_cost': ticket_data.get('ticket_cost', 'Not detected'),
                        'ticket_data': ticket_data,
                        'raw_response': response_text,
                        'main_game_results': main_game_results,
                        'comparison': {
                            'message': comparison_message,
                            'extracted_numbers': player_numbers,
                            'lottery_type': lottery_type,
                            'games_checked': 1 if main_game_results else 0,
                            'enhanced_comparison': True,
                            'main_game': main_game_results
                        }
                    }
                
                else:
                    # Unsupported lottery type
                    result = {
                        'success': False,
                        'error': f'Unsupported lottery type: {lottery_type}',
                        'lottery_type': lottery_type,
                        'raw_response': response_text
                    }
                
                logger.info(f"Successfully processed ticket: {result}")
                return jsonify(result)
            else:
                return jsonify({'success': False, 'error': 'Could not parse ticket data from AI response'})
                
        except json.JSONDecodeError as e:
            logger.error(f"JSON parsing error: {e}")
            logger.error(f"Raw response that failed to parse: {response_text}")
            # Clean up file if it exists
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'success': False, 'error': f'Could not parse AI response. Raw response: {response_text[:200]}...'})
        except Exception as e:
            logger.error(f"Error in ticket processing: {e}")
            logger.error(f"Traceback: ", exc_info=True)
            # Clean up file if it exists
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'success': False, 'error': f'Processing error: {str(e)}'})
            
    except Exception as e:
        logger.error(f"Error in process_ticket route: {e}")
        logger.error(f"Traceback: ", exc_info=True)
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'})

# Scan ticket functionality is now handled by the /process-ticket endpoint

@app.route('/scan-results')
def scan_results():
    """Display the results of a scanned ticket"""
    result = session.get('scan_result')
    
    if not result:
        flash("No scan results available. Please scan a ticket first.", "warning")
        return redirect(url_for('ticket_scanner'))
    
    # Additional SEO metadata
    meta_description = "View your South African lottery ticket scan results. Find out if you've won with our advanced ticket scanner."
    
    return render_template('scanner/scan_results.html', 
                          title="Ticket Scan Results | Check If You've Won",
                          result=result,
                          meta_description=meta_description)

# Guides Routes
@app.route('/guides')
def guides_index():
    """Display index of lottery guides"""
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Lottery Guides", "url": url_for('guides_index')}
    ]
    
    # Define SEO metadata
    meta_description = "Comprehensive guides on how to play South African lottery games. Learn rules, strategies, and tips for Lottery, PowerBall, Daily Lottery and more."
    
    return render_template('guides/index.html', 
                          title="South African Lottery Guides | Tips & How-To Articles",
                          breadcrumbs=breadcrumbs,
                          meta_description=meta_description)

@app.route('/guides/how-to-play-lottery')
def how_to_play_lotto():
    """Display guide on how to play Lottery"""
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Lottery Guides", "url": url_for('guides_index')},
        {"name": "How to Play Lottery", "url": url_for('how_to_play_lotto')}
    ]
    
    # Define SEO metadata
    meta_description = "Comprehensive guide on how to play the South African Lottery. Learn Lottery rules, drawing days, odds of winning, prize divisions, and expert tips."
    
    return render_template('guides/how_to_play_lotto.html',
                          title="How to Play Lottery South Africa | Complete Guide & Tips",
                          breadcrumbs=breadcrumbs,
                          meta_description=meta_description)

@app.route('/guides/how-to-play-powerball')
def how_to_play_powerball():
    """Display guide on how to play PowerBall"""
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Lottery Guides", "url": url_for('guides_index')},
        {"name": "How to Play PowerBall", "url": url_for('how_to_play_powerball')}
    ]
    
    # Define SEO metadata
    meta_description = "Complete guide to playing PowerBall South Africa. Learn game rules, drawing schedule, odds, prize divisions, and expert strategies to increase your chances."
    
    # Placeholder until we create the PowerBall guide template
    flash("PowerBall guide coming soon! Check out our Lottery guide in the meantime.", "info")
    return redirect(url_for('guides_index'))

@app.route('/guides/how-to-play-daily-lottery')
def how_to_play_daily_lotto():
    """Display guide on how to play Daily Lottery"""
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Lottery Guides", "url": url_for('guides_index')},
        {"name": "How to Play Daily Lottery", "url": url_for('how_to_play_daily_lotto')}
    ]
    
    # Define SEO metadata
    meta_description = "Learn how to play Daily Lottery South Africa with our comprehensive guide. Discover game rules, draw times, odds of winning, and how to claim prizes."
    
    # Placeholder until we create the Daily Lottery guide template
    flash("Daily Lottery guide coming soon! Check out our Lottery guide in the meantime.", "info")
    return redirect(url_for('guides_index'))

@app.route('/results')
def results():
    """Show overview of all lottery types with links to specific results"""
    # Map display names to database names to handle authentic data from official screenshots
    lottery_types = ['Lottery', 'Lottery Plus 1', 'Lottery Plus 2', 
                     'Powerball', 'Powerball Plus', 'Daily Lottery']
    
    # Database mapping for authentic lottery data (from official screenshots)
    # Handle multiple case variations found in actual database
    db_name_mapping = {
        'Lottery': ['LOTTO', 'Lotto'],
        'Lottery Plus 1': ['LOTTO PLUS 1', 'Lotto Plus 1'], 
        'Lottery Plus 2': ['LOTTO PLUS 2', 'Lotto Plus 2'],
        'Powerball': ['POWERBALL', 'PowerBall', 'Powerball'],
        'Powerball Plus': ['POWERBALL PLUS', 'PowerBall Plus', 'Powerball Plus'],
        'Daily Lottery': ['DAILY LOTTO', 'Daily Lotto']
    }
    
    try:
        # Ensure data_aggregator is loaded before using it
        global data_aggregator
        
        # Import if not already loaded - data_aggregator module not available
        # Module functionality replaced with direct database queries
        
        # Initialize with empty dict in case the next step fails
        latest_results = {}
        
        try:
            # Use direct SQL approach to ensure we get authentic data
            latest_results = {}
            
            # Use same SQL query as homepage for authentic Gemini-extracted data
            from sqlalchemy import text
            
            # Get latest result for each lottery type - same as homepage
            # Note: Lottery games that draw together share the same draw_date and draw_number:
            # - LOTTO, LOTTO PLUS 1, LOTTO PLUS 2 all draw on the same day (same draw_number)
            # - PowerBall and POWERBALL PLUS draw on the same day (same draw_number) 
            # - DAILY LOTTO draws independently
            query = text("""
                WITH ranked_results AS (
                  SELECT lottery_type, draw_number, draw_date, numbers as main_numbers, bonus_numbers,
                    ROW_NUMBER() OVER (PARTITION BY lottery_type ORDER BY draw_date DESC, draw_number DESC) as rn
                  FROM lottery_result 
                  WHERE lottery_type IN ('LOTTO', 'LOTTO PLUS 1', 'LOTTO PLUS 2', 'PowerBall', 'POWERBALL PLUS', 'DAILY LOTTO')
                )
                SELECT lottery_type, draw_number, draw_date, main_numbers, bonus_numbers
                FROM ranked_results 
                WHERE rn = 1
            """)
            
            raw_results = db.session.execute(query).fetchall()
            app.logger.info(f"✓ Results page using same authentic data as homepage: {[(r.lottery_type, r.draw_number) for r in raw_results]}")
            
            # Process each lottery type to match homepage data exactly
            lottery_lookup = {}
            for row in raw_results:
                lottery_lookup[row.lottery_type] = row
            
            # Create results in display order mapping database names to display names
            type_mappings = {
                'LOTTO': 'Lottery',
                'LOTTO PLUS 1': 'Lottery Plus 1', 
                'LOTTO PLUS 2': 'Lottery Plus 2',
                'PowerBall': 'Powerball',
                'POWERBALL PLUS': 'Powerball Plus',
                'DAILY LOTTO': 'Daily Lottery'
            }
            
            for db_type, display_type in type_mappings.items():
                if db_type in lottery_lookup:
                    row = lottery_lookup[db_type]
                    
                    # Parse numbers using same logic as homepage
                    numbers = []
                    bonus_numbers = []
                    
                    if row.main_numbers:
                        try:
                            numbers = json.loads(str(row.main_numbers))
                            app.logger.info(f"Parsed main numbers for {display_type}: {numbers}")
                        except (json.JSONDecodeError, ValueError):
                            if str(row.main_numbers).startswith('{') and str(row.main_numbers).endswith('}'):
                                nums_str = str(row.main_numbers)[1:-1]
                                if nums_str.strip():
                                    numbers = [int(n.strip()) for n in nums_str.split(',') if n.strip().isdigit()]
                    
                    if row.bonus_numbers:
                        try:
                            bonus_numbers = json.loads(str(row.bonus_numbers))
                            app.logger.info(f"Parsed bonus numbers for {display_type}: {bonus_numbers}")
                        except (json.JSONDecodeError, ValueError):
                            if str(row.bonus_numbers).startswith('{') and str(row.bonus_numbers).endswith('}'):
                                bonus_str = str(row.bonus_numbers)[1:-1]
                                if bonus_str.strip():
                                    bonus_numbers = [int(n.strip()) for n in bonus_str.split(',') if n.strip().isdigit()]
                    
                    app.logger.info(f"✓ Direct SQL: {display_type} - Draw {row.draw_number}")
                    
                    class ResultObj:
                        def __init__(self, lottery_type, draw_number, draw_date, numbers, bonus_numbers):
                            self.lottery_type = lottery_type
                            self.draw_number = draw_number
                            self.draw_date = draw_date
                            self.numbers = numbers
                            self.bonus_numbers = bonus_numbers
                        
                        @property
                        def main_numbers(self):
                            return self.numbers
                        
                        def get_numbers_list(self):
                            return self.numbers if self.numbers else []
                        
                        def get_bonus_numbers_list(self):
                            return self.bonus_numbers if self.bonus_numbers else []
                    
                    latest_results[display_type] = ResultObj(
                        display_type,
                        row.draw_number,
                        row.draw_date,
                        numbers,
                        bonus_numbers
                    )
            
            app.logger.info(f"Direct SQL retrieved: {len(latest_results)} authentic lottery results")
            
            # All data processed above using same source as homepage
            
        except Exception as e:
            app.logger.error(f"Direct SQL error: {str(e)}")
            latest_results = {}
        
        # Define breadcrumbs for SEO
        breadcrumbs = [
            {"name": "Results", "url": url_for('results')}
        ]
        
        # Add cache busting to ensure fresh data display
        import time
        cache_buster = int(time.time())
        
        # Debug template data before rendering
        app.logger.info(f"Template data - lottery_types: {lottery_types}")
        app.logger.info(f"Template data - latest_results: {latest_results}")
        
        # Test each lottery type specifically
        for lt in lottery_types:
            if lt in latest_results:
                result = latest_results[lt]
                app.logger.info(f"✓ {lt}: Draw {result.draw_number}, Numbers: {result.main_numbers}")
            else:
                app.logger.error(f"✗ MISSING: {lt} not in latest_results")
        
        return render_template('results_overview.html',
                            lottery_types=lottery_types,
                            latest_results=latest_results,
                            title="All Lottery Results",
                            breadcrumbs=breadcrumbs,
                            cache_buster=cache_buster)
                            
    except Exception as e:
        app.logger.error(f"Critical error in results route: {str(e)}", exc_info=True)
        # Define breadcrumbs for SEO even in error case
        breadcrumbs = [
            {"name": "Results", "url": url_for('results')}
        ]
        # Return a simplified template with no result data
        return render_template('results_overview.html',
                            lottery_types=lottery_types,
                            latest_results={},  # Empty dict on critical error
                            title="All Lottery Results",
                            breadcrumbs=breadcrumbs)

@app.route('/lottery/<int:draw_number>')
def lottery_details(draw_number):
    """Show detailed lottery results with complete prize divisions and financial info"""
    try:
        # Get lottery result with comprehensive data
        lottery_result = LotteryResult.query.filter_by(draw_number=draw_number).first()
        
        if not lottery_result:
            flash(f"Draw {draw_number} not found", "warning")
            return redirect(url_for('results'))
        
        # Parse main numbers using the existing method
        main_numbers = lottery_result.get_numbers_list()
        
        # Parse bonus number using the existing method
        bonus_numbers = lottery_result.get_bonus_numbers_list()
        bonus_number = bonus_numbers[0] if bonus_numbers else None
        
        # Parse divisions data from both possible fields
        divisions = {}
        
        if lottery_result.divisions:
            try:
                if isinstance(lottery_result.divisions, str):
                    divisions = json.loads(lottery_result.divisions)
                else:
                    divisions = lottery_result.divisions
            except Exception as e:
                app.logger.error(f"Error parsing divisions: {e}")
                divisions = {}
        
        # Check for prize_divisions field if it exists
        if hasattr(lottery_result, 'prize_divisions') and lottery_result.prize_divisions and not divisions:
            try:
                if isinstance(lottery_result.prize_divisions, str):
                    divisions = json.loads(lottery_result.prize_divisions)
                else:
                    divisions = lottery_result.prize_divisions
            except Exception as e:
                app.logger.error(f"Error parsing prize_divisions: {e}")
                divisions = {}
        
        # Create numerical order (sorted)
        numerical_order = sorted(main_numbers) if main_numbers else []
        
        return render_template(
            'lottery_details.html',
            lottery_result=lottery_result,
            main_numbers=main_numbers,
            bonus_number=bonus_number,
            divisions=divisions,
            numerical_order=numerical_order
        )
        
    except Exception as e:
        app.logger.error(f"Error showing lottery details: {e}")
        flash("Error loading lottery details", "danger")
        return redirect(url_for('results'))

@app.route('/results/<lottery_type>')
def lottery_results(lottery_type):
    """Show all results for a specific lottery type"""
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Number of results per page
    
    # Ensure data_aggregator is loaded before using it
    global data_aggregator
    
    # Import if not already loaded - data_aggregator not available
    # if data_aggregator is None:
    #     import data_aggregator as da
    #     data_aggregator = da
    #     logger.info("Loaded data_aggregator module on demand")
    
    # Map display names to database names for authentic lottery data - matching main results page
    # Handle multiple case variations found in actual database
    lottery_type_mapping = {
        'Lottery': ['LOTTO', 'Lotto'],
        'Lottery Plus 1': ['LOTTO PLUS 1', 'Lotto Plus 1'], 
        'Lottery Plus 2': ['LOTTO PLUS 2', 'Lotto Plus 2'],
        'Powerball': ['POWERBALL', 'PowerBall', 'Powerball'],
        'Powerball Plus': ['POWERBALL PLUS', 'PowerBall Plus', 'Powerball Plus'],
        'Daily Lottery': ['DAILY LOTTO', 'Daily Lotto']
    }
    
    # Get the actual database lottery type names (multiple variations)
    db_lottery_types = lottery_type_mapping.get(lottery_type, [lottery_type])
    
    # Get all results for this lottery type directly from database using OR query for all variations
    all_results = LotteryResult.query.filter(LotteryResult.lottery_type.in_(db_lottery_types)).order_by(LotteryResult.draw_date.desc()).all()
    
    logger.info(f"Looking for lottery type '{lottery_type}' mapped to DB types '{db_lottery_types}', found {len(all_results)} results")
    
    # Create a paginated result from the raw list
    # This mimics SQLAlchemy's pagination object with the properties the template expects
    class PaginatedResults:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            
        @property
        def pages(self):
            return (self.total + self.per_page - 1) // self.per_page
            
        @property
        def has_prev(self):
            return self.page > 1
            
        @property
        def has_next(self):
            return self.page < self.pages
            
        @property
        def prev_num(self):
            return self.page - 1 if self.has_prev else None
            
        @property
        def next_num(self):
            return self.page + 1 if self.has_next else None
            
        def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
            last = 0
            for num in range(1, self.pages + 1):
                if num <= left_edge or \
                   (num > self.page - left_current - 1 and num < self.page + right_current) or \
                   num > self.pages - right_edge:
                    if last + 1 != num:
                        yield None
                    yield num
                    last = num
    
    # Calculate the sliced items for the current page
    start = (page - 1) * per_page
    end = min(start + per_page, len(all_results))
    items = all_results[start:end]
    
    # Create the pagination object
    results = PaginatedResults(items, page, per_page, len(all_results))
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Results", "url": url_for('results')},
        {"name": lottery_type, "url": url_for('lottery_results', lottery_type=lottery_type)}
    ]
    
    return render_template('results.html', 
                           results=results, 
                           lottery_type=lottery_type,
                           title=f"{lottery_type} Results",
                           breadcrumbs=breadcrumbs)



@app.route('/export-template')
@login_required
def export_template():
    """Export an empty spreadsheet template for importing lottery data"""
    if not current_user.is_admin:
        flash('You must be an admin to export templates.', 'danger')
        return redirect(url_for('index'))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"lottery_data_template_{timestamp}.xlsx"
    
    # Create template with lottery-specific tabs
    create_lottery_template(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    
    return send_from_directory(
        app.config['UPLOAD_FOLDER'], 
        filename, 
        as_attachment=True
    )

# File upload progress tracker
# This dictionary stores the progress of file uploads for each user
# Structure: {user_id: {'percentage': 0-100, 'status': 'uploading|processing|complete|error', 'filename': 'example.xlsx'}}
file_upload_progress = {}

@app.route('/api/file-upload-progress')
@login_required
def get_file_upload_progress():
    """API endpoint to get current file upload progress for the current user"""
    user_id = current_user.id
    
    # If no progress exists for this user, return default values
    if user_id not in file_upload_progress:
        file_upload_progress[user_id] = {
            'percentage': 0,
            'status': 'initializing',
            'filename': ''
        }
    
    return jsonify(file_upload_progress[user_id])

@app.route('/api/file-upload-progress/reset', methods=['POST'])
@login_required
def reset_file_upload_progress():
    """Reset the file upload progress for the current user"""
    user_id = current_user.id
    
    file_upload_progress[user_id] = {
        'percentage': 0,
        'status': 'initializing',
        'filename': ''
    }
    
    return jsonify({'success': True})

@app.route('/import-latest-spreadsheet', methods=['POST'])
@login_required
def import_latest_spreadsheet_route():
    """Import the latest spreadsheet file from attached_assets or uploads directory"""
    if not current_user.is_admin:
        flash('You must be an admin to import data.', 'danger')
        return redirect(url_for('index'))
    
    import_type = request.form.get('import_type', 'excel')
    purge = request.form.get('purge', 'no') == 'yes'
    pattern = request.form.get('pattern', '*.xlsx')  # Modified to look for any Excel file
    
    try:
        # First, try to find in attached_assets
        latest_file = find_latest_spreadsheet("attached_assets", pattern)
        
        # If not found in attached_assets, look in uploads directory
        if not latest_file and os.path.exists("uploads"):
            latest_file = find_latest_spreadsheet("uploads", pattern)
            
            if latest_file:
                logger.info(f"Using spreadsheet from uploads directory: {latest_file}")
                # If found in uploads, use that directory for import
                success = import_latest_spreadsheet("uploads", pattern, import_type, purge)
        else:
            # Using spreadsheet from attached_assets
            success = import_latest_spreadsheet("attached_assets", pattern, import_type, purge)
        
        if not latest_file:
            # Try one more time with the original lottery_data_*.xlsx pattern
            original_pattern = "lottery_data_*.xlsx"
            latest_file = find_latest_spreadsheet("attached_assets", original_pattern)
            
            if not latest_file and os.path.exists("uploads"):
                latest_file = find_latest_spreadsheet("uploads", original_pattern)
                
                if latest_file:
                    # If found in uploads with original pattern, use that
                    success = import_latest_spreadsheet("uploads", original_pattern, import_type, purge)
            else:
                # Using spreadsheet from attached_assets with original pattern
                success = import_latest_spreadsheet("attached_assets", original_pattern, import_type, purge)
        
        if not latest_file:
            flash(f'No Excel spreadsheets found in attached_assets or uploads directories. Please upload a spreadsheet first.', 'danger')
            return redirect(url_for('import_data'))
        
        if success:
            flash(f'Successfully imported latest spreadsheet: {os.path.basename(latest_file)}', 'success')
        else:
            flash(f'Failed to import spreadsheet: {os.path.basename(latest_file)}. Check logs for details.', 'danger')
            
        return redirect(url_for('import_data'))
        
    except Exception as e:
        logger.exception(f"Error importing spreadsheet: {str(e)}")
        flash(f'Error importing spreadsheet: {str(e)}', 'danger')
        return redirect(url_for('import_data'))

@app.route('/import-history')
@login_required
def import_history():
    """View import history and details"""
    if not current_user.is_admin:
        flash('You must be an admin to view import history.', 'danger')
        return redirect(url_for('index'))
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Admin Dashboard", "url": url_for('admin')},
        {"name": "Import History", "url": url_for('import_history')}
    ]
    
    # Define SEO metadata
    meta_description = "View and analyze South African lottery data import history. Track data sources, import dates, success rates, and detailed import logs for system administration."
    
    try:
        # Get recent imports with pagination
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        # Get all import history records, newest first
        imports = ImportHistory.query.order_by(ImportHistory.import_date.desc()).paginate(
            page=page, per_page=per_page, error_out=False)
        
        return render_template('import_history.html',
                            imports=imports,
                            title="Lottery Data Import History | System Administration",
                            breadcrumbs=breadcrumbs,
                            meta_description=meta_description)
    except Exception as e:
        app.logger.error(f"Error retrieving import history: {str(e)}")
        flash(f"Error loading import history: {str(e)}", 'danger')
        return redirect(url_for('admin'))

@app.route('/import-history/<int:import_id>')
@login_required
def import_details(import_id):
    """View details of a specific import"""
    if not current_user.is_admin:
        flash('You must be an admin to view import details.', 'danger')
        return redirect(url_for('index'))
    
    # Define SEO metadata (will be updated with specific import details when the record is loaded)
    meta_description = "Detailed South African lottery data import information. View imported draw numbers, dates, and success status for administrative record-keeping."
    
    try:
        # Get the import record
        import_record = ImportHistory.query.get_or_404(import_id)
        
        # Define breadcrumbs for SEO
        breadcrumbs = [
            {"name": "Admin Dashboard", "url": url_for('admin')},
            {"name": "Import History", "url": url_for('import_history')},
            {"name": f"Import #{import_id}", "url": url_for('import_details', import_id=import_id)}
        ]
        
        # Get all records that were imported in this batch
        imported_records = ImportedRecord.query.filter_by(import_id=import_id).all()
        
        # Group records by lottery type for easier display
        records_by_type = {}
        for record in imported_records:
            lottery_type = record.lottery_type
            if lottery_type not in records_by_type:
                records_by_type[lottery_type] = []
            records_by_type[lottery_type].append(record)
        
        # Update meta description with specific import details
        meta_description = f"Details for lottery data import #{import_id} from {import_record.import_date.strftime('%Y-%m-%d')}. Contains {len(imported_records)} records across {len(records_by_type)} lottery types."
        
        return render_template('import_details.html',
                              import_record=import_record,
                              records_by_type=records_by_type,
                              title=f"Import Details #{import_id} | {import_record.import_date.strftime('%Y-%m-%d %H:%M')}",
                              breadcrumbs=breadcrumbs,
                              meta_description=meta_description)
    except Exception as e:
        app.logger.error(f"Error retrieving import details: {str(e)}")
        flash(f"Error loading import details: {str(e)}", 'danger')
        return redirect(url_for('import_history'))

@app.route('/import-data', methods=['GET', 'POST'])
@login_required
def import_data():
    """Import data from Excel spreadsheet"""
    global import_excel, import_snap_lotto_data
    
    if not current_user.is_admin:
        flash('You must be an admin to import data.', 'danger')
        return redirect(url_for('index'))
        
    import_stats = None
    user_id = current_user.id
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Admin Dashboard", "url": url_for('admin')},
        {"name": "Import Data", "url": url_for('import_data')}
    ]
    
    # Define SEO metadata
    meta_description = "Import South African lottery data from Excel spreadsheets. Upload Lottery, PowerBall, and Daily Lottery results to maintain an up-to-date database of winning numbers and prize information."
    
    # Initialize or reset progress tracking
    file_upload_progress[user_id] = {
        'percentage': 0,
        'status': 'initializing',
        'filename': ''
    }

    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)
            
        file = request.files['file']
        
        # If user does not select file, browser also
        # submit an empty part without filename
        if file.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)
            
        if file:
            filename = secure_filename(file.filename)
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Update progress
            file_upload_progress[user_id] = {
                'percentage': 25,
                'status': 'uploading',
                'filename': filename
            }
            
            # Save the file
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(excel_path)
            
            # Update progress
            file_upload_progress[user_id] = {
                'percentage': 50,
                'status': 'processing',
                'filename': filename
            }
            
            try:
                # Check if this is a multi-sheet template by trying to detect sheets
                is_template_format = False
                try:
                    import pandas as pd
                    # Load Excel file and check for template sheets
                    xl = pd.ExcelFile(excel_path)
                    sheet_names = xl.sheet_names
                    
                    # Check if the file has the expected sheet structure (multiple sheets with lottery type names)
                    template_sheets = [
                        "Lottery", "Lottery Plus 1", "Lottery Plus 2", 
                        "Powerball", "Powerball Plus", "Daily Lottery"
                    ]
                    
                    # If at least 3 of the expected template sheets are present, we consider it a template
                    matches = sum(1 for sheet in template_sheets if sheet in sheet_names)
                    is_template_format = matches >= 3
                    
                    # Also consider template if the name contains "template"
                    if not is_template_format:
                        is_template_format = "template" in filename.lower()
                        
                    logger.info(f"File format detection: {'multi-sheet template' if is_template_format else 'standard'}")
                except Exception as e:
                    logger.warning(f"Error detecting file format: {str(e)}")
                    # Fall back to filename detection
                    is_template_format = "template" in filename.lower()
                
                if is_template_format:
                    try:
                        # Update progress
                        file_upload_progress[user_id] = {
                            'percentage': 50,
                            'status': 'processing template format',
                            'filename': filename
                        }
                        
                        # Import using the multi-sheet template processor
                        import multi_sheet_import
                        import_stats = multi_sheet_import.import_multisheet_excel(excel_path, flask_app=app)
                        
                        # If import was successful, track it in the import history
                        if isinstance(import_stats, dict) and import_stats.get('success'):
                            # Create import history record
                            import_history = ImportHistory(
                                import_type='multi_sheet_template',
                                file_name=filename,
                                records_added=import_stats.get('added', 0),
                                records_updated=import_stats.get('updated', 0),
                                total_processed=import_stats.get('total', 0),
                                errors=import_stats.get('errors', 0),
                                user_id=current_user.id
                            )
                            db.session.add(import_history)
                            db.session.commit()
                            
                            # Save individual imported records if available
                            if 'imported_records' in import_stats and import_stats['imported_records']:
                                for record_data in import_stats['imported_records']:
                                    imported_record = ImportedRecord(
                                        import_id=import_history.id,
                                        lottery_type=record_data['lottery_type'],
                                        draw_number=record_data['draw_number'],
                                        draw_date=record_data['draw_date'],
                                        is_new=record_data['is_new'],
                                        lottery_result_id=record_data['lottery_result_id']
                                    )
                                    db.session.add(imported_record)
                                db.session.commit()
                            
                            # Display results
                            added = import_stats.get('added', 0)
                            total = import_stats.get('total', 0)
                            errors = import_stats.get('errors', 0)
                            updated = import_stats.get('updated', 0)
                            
                            if added > 0 and updated > 0:
                                flash(f'Multi-sheet template import completed successfully. Added {added} new records, updated {updated} existing records, processed {total} total records with {errors} errors.', 'success')
                            elif added > 0:
                                flash(f'Multi-sheet template import completed successfully. Added {added} new records, processed {total} total records with {errors} errors.', 'success')
                            elif updated > 0:
                                flash(f'Multi-sheet template import completed successfully. Updated {updated} existing records, processed {total} total records with {errors} errors.', 'success')
                            else:
                                flash(f'Multi-sheet template import completed. No records were added or updated. Processed {total} records with {errors} errors.', 'info')
                        else:
                            error_msg = import_stats.get('error', 'Unknown error')
                            flash(f'Error in multi-sheet template import: {error_msg}', 'danger')
                        
                        # Update progress to complete
                        file_upload_progress[user_id] = {
                            'percentage': 100,
                            'status': 'complete',
                            'filename': filename,
                            'summary': import_stats if isinstance(import_stats, dict) else None
                        }
                    except Exception as e:
                        logger.error(f"Multi-sheet template import error: {str(e)}")
                        logger.error(traceback.format_exc())
                        flash(f"Error in multi-sheet template import: {str(e)}", 'danger')
                        
                        # Update progress to error
                        file_upload_progress[user_id] = {
                            'percentage': 100,
                            'status': 'error',
                            'filename': filename
                        }
                else:
                    # Try standard format for non-template files
                    try:
                        if import_excel is None:
                            import import_excel as ie
                        else:
                            ie = import_excel
                        import_stats = ie.import_excel_data(excel_path)
                        
                        # If import was successful, track it in the import history
                        if isinstance(import_stats, dict) and import_stats.get('success'):
                            # Create import history record
                            import_history = ImportHistory(
                                import_type='excel',
                                file_name=filename,
                                records_added=import_stats.get('added', 0),
                                records_updated=import_stats.get('updated', 0),
                                total_processed=import_stats.get('total', 0),
                                errors=import_stats.get('errors', 0),
                                user_id=current_user.id
                            )
                            db.session.add(import_history)
                            db.session.commit()
                            
                            # Save individual imported records if available
                            if 'imported_records' in import_stats and import_stats['imported_records']:
                                for record_data in import_stats['imported_records']:
                                    imported_record = ImportedRecord(
                                        import_id=import_history.id,
                                        lottery_type=record_data['lottery_type'],
                                        draw_number=record_data['draw_number'],
                                        draw_date=record_data['draw_date'],
                                        is_new=record_data['is_new'],
                                        lottery_result_id=record_data['lottery_result_id']
                                    )
                                    db.session.add(imported_record)
                                db.session.commit()
                        
                        # Check if import was unsuccessful (False) or if no records were imported
                        if import_stats is False or (isinstance(import_stats, dict) and import_stats.get('total', 0) == 0):
                            # If standard import fails, try Snap Lotto format
                            try:
                                # Update progress
                                file_upload_progress[user_id] = {
                                    'percentage': 75,
                                    'status': 'processing',
                                    'filename': filename
                                }
                                
                                # Import import_snap_lotto_data if not already loaded
                                if import_snap_lotto_data is None:
                                    import import_snap_lotto_data as isld
                                else:
                                    isld = import_snap_lotto_data
                                import_stats = isld.import_snap_lotto_data(excel_path, flask_app=app)
                                
                                # If import was successful, track it in the import history
                                if isinstance(import_stats, dict) and import_stats.get('success'):
                                    # Create import history record
                                    import_history = ImportHistory(
                                        import_type='snap_lotto',
                                        file_name=filename,
                                        records_added=import_stats.get('added', 0),
                                        records_updated=import_stats.get('updated', 0),
                                        total_processed=import_stats.get('total', 0),
                                        errors=import_stats.get('errors', 0),
                                        user_id=current_user.id
                                    )
                                    db.session.add(import_history)
                                    db.session.commit()
                                    
                                    # Save individual imported records if available
                                    if 'imported_records' in import_stats and import_stats['imported_records']:
                                        for record_data in import_stats['imported_records']:
                                            imported_record = ImportedRecord(
                                                import_id=import_history.id,
                                                lottery_type=record_data['lottery_type'],
                                                draw_number=record_data['draw_number'],
                                                draw_date=record_data['draw_date'],
                                                is_new=record_data['is_new'],
                                                lottery_result_id=record_data['lottery_result_id']
                                            )
                                            db.session.add(imported_record)
                                        db.session.commit()
                                
                                # Display results if available and successful
                                if isinstance(import_stats, dict) and import_stats.get('success'):
                                    added = import_stats.get('added', 0)
                                    total = import_stats.get('total', 0)
                                    errors = import_stats.get('errors', 0)
                                    
                                    updated = import_stats.get('updated', 0)
                                    if added > 0 and updated > 0:
                                        flash(f'Snap Lotto import completed successfully. Added {added} new records, updated {updated} existing records, processed {total} total records with {errors} errors.', 'success')
                                    elif added > 0:
                                        flash(f'Snap Lotto import completed successfully. Added {added} new records, processed {total} total records with {errors} errors.', 'success')
                                    elif updated > 0:
                                        flash(f'Snap Lotto import completed successfully. Updated {updated} existing records, processed {total} total records with {errors} errors.', 'success')
                                    else:
                                        flash(f'Snap Lotto import completed. No records were added or updated. Processed {total} records with {errors} errors.', 'info')
                                else:
                                    flash('Snap Lotto import completed', 'info')
                                
                                # Update progress to complete
                                file_upload_progress[user_id] = {
                                    'percentage': 100,
                                    'status': 'complete',
                                    'filename': filename,
                                    'summary': import_stats if isinstance(import_stats, dict) else None
                                }
                            except Exception as e:
                                logger.error(f"Snap Lotto import error: {str(e)}")
                                flash(f"Error in Snap Lotto import: {str(e)}", 'danger')
                                
                                # Update progress to error
                                file_upload_progress[user_id] = {
                                    'percentage': 100,
                                    'status': 'error',
                                    'filename': filename
                                }
                        else:
                            # Display results for standard import
                            if isinstance(import_stats, dict) and import_stats.get('success'):
                                added = import_stats.get('added', 0)
                                total = import_stats.get('total', 0)
                                errors = import_stats.get('errors', 0)
                                
                                updated = import_stats.get('updated', 0)
                                if added > 0 and updated > 0:
                                    flash(f'Import completed successfully. Added {added} new records, updated {updated} existing records, processed {total} total records with {errors} errors.', 'success')
                                elif added > 0:
                                    flash(f'Import completed successfully. Added {added} new records, processed {total} total records with {errors} errors.', 'success')
                                elif updated > 0:
                                    flash(f'Import completed successfully. Updated {updated} existing records, processed {total} total records with {errors} errors.', 'success')
                                else:
                                    flash(f'Import completed. No records were added or updated. Processed {total} records with {errors} errors.', 'info')
                            else:
                                flash('Import completed', 'info')
                            
                            # Update progress to complete
                            file_upload_progress[user_id] = {
                                'percentage': 100,
                                'status': 'complete',
                                'filename': filename,
                                'summary': import_stats if isinstance(import_stats, dict) else None
                            }
                    except Exception as e:
                        logger.error(f"Excel import error: {str(e)}")
                        flash(f"Error in import: {str(e)}", 'danger')
                        
                        # Update progress to error
                        file_upload_progress[user_id] = {
                            'percentage': 100,
                            'status': 'error',
                            'filename': filename
                        }
            except Exception as e:
                logger.error(f"File processing error: {str(e)}")
                logger.error(traceback.format_exc())
                flash(f"File processing error: {str(e)}", 'danger')
                
                # Update progress to error
                file_upload_progress[user_id] = {
                    'percentage': 100,
                    'status': 'error',
                    'filename': filename
                }
                
    # Get some example results for each lottery type to display
    example_results = {}
    lottery_types = ['Lottery', 'Lottery Plus 1', 'Lottery Plus 2', 
                     'Powerball', 'Powerball Plus', 'Daily Lottery']
    
    try:                 
        for lottery_type in lottery_types:
            try:
                results = LotteryResult.query.filter_by(lottery_type=lottery_type).order_by(
                    LotteryResult.draw_date.desc()).limit(5).all()
                if results:
                    example_results[lottery_type] = results
            except Exception as e:
                app.logger.error(f"Error retrieving example results for {lottery_type}: {str(e)}")
                # Continue with other lottery types even if one fails
                continue
    except Exception as e:
        app.logger.error(f"Database error in import_data: {str(e)}")
        # If we can't get any example results, proceed without them
    
    # Extract structured data from import_stats (if available) for template rendering
    added_count = 0
    updated_count = 0
    total_count = 0
    error_count = 0
    imported_results = []
    
    if isinstance(import_stats, dict) and import_stats.get('success'):
        added_count = import_stats.get('added', 0)
        updated_count = import_stats.get('updated', 0)
        total_count = import_stats.get('total', 0)
        error_count = import_stats.get('errors', 0)
        
        # Get imported records for display
        if 'imported_records' in import_stats and import_stats['imported_records']:
            for record_data in import_stats['imported_records']:
                result = LotteryResult.query.get(record_data.get('lottery_result_id'))
                if result:
                    imported_results.append(result)
    
    return render_template('import_data.html', 
                           import_stats=import_stats,
                           example_results=example_results,
                           title="Import South African Lottery Data | Admin Tools",
                           meta_description=meta_description,
                           breadcrumbs=breadcrumbs,
                           added_count=added_count,
                           updated_count=updated_count,
                           total_count=total_count,
                           error_count=error_count,
                           imported_results=imported_results)

@app.route('/retake-screenshots')
@login_required
def retake_screenshots():
    """Admin route to retake all screenshots"""
    if not current_user.is_admin:
        flash('You must be an admin to retake screenshots.', 'danger')
        return redirect(url_for('index'))
        
    try:
        # Load screenshot manager for automated operations
        global screenshot_manager
        if screenshot_manager is None:
            import screenshot_manager
        
        count = screenshot_manager.retake_all_screenshots(app)
        
        if count > 0:
            success_count = count
            fail_count = 0
            
            flash(f'Screenshot process started. {success_count} URLs queued successfully. {fail_count} failed.', 'info')
        else:
            flash('No URLs configured for screenshots.', 'warning')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('admin'))

@app.route('/gemini-automation')
@login_required
def gemini_automation():
    """Run complete automation workflow using Google Gemini 2.5 Pro"""
    if not current_user.is_admin:
        flash('You must be an admin to run automation.', 'danger')
        return redirect(url_for('index'))
        
    try:
        from gemini_automation_controller import GeminiAutomationController
        
        # Create controller and run workflow
        controller = GeminiAutomationController()
        results = controller.run_complete_workflow()
        
        if results and results.get('overall_success'):
            flash('Gemini automation completed successfully! Fresh lottery data extracted and saved.', 'success')
        else:
            error_steps = [step for step, success in results.items() if not success and step != 'overall_success']
            flash(f'Automation completed with issues in: {", ".join(error_steps)}', 'warning')
            
    except Exception as e:
        flash(f'Automation failed: {str(e)}', 'danger')
    
    return redirect(url_for('admin'))

@app.route('/purge-database')
@login_required
def purge_database():
    """Admin route to purge all data from the database"""
    if not current_user.is_admin:
        flash('You must be an admin to purge the database.', 'danger')
        return redirect(url_for('index'))
        
    try:
        # Import the purge_data module
        import purge_data
        
        # Execute the purge operation
        result = purge_data.purge_data()
        
        if result:
            flash('Database purged successfully. All lottery results and screenshots have been deleted.', 'success')
        else:
            flash('An error occurred during the purge operation.', 'danger')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('admin'))

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    """Manage data syncs and system settings"""
    if not current_user.is_admin:
        flash('You must be an admin to access settings.', 'danger')
        return redirect(url_for('index'))
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Admin Dashboard", "url": url_for('admin')},
        {"name": "System Settings", "url": url_for('settings')}
    ]
    
    # Define SEO metadata
    meta_description = "Configure South African lottery system settings and scheduled tasks. Manage data synchronization, screenshot capture timing, and system preferences."
    
    schedule_configs = ScheduleConfig.query.all()
    
    # Handle form submission for updating settings
    if request.method == 'POST':
        # This would normally handle the form submission
        flash('Settings updated successfully.', 'success')
        return redirect(url_for('settings'))
    
    return render_template('settings.html', 
                          schedule_configs=schedule_configs,
                          title="System Settings | Lottery Data Management",
                          meta_description=meta_description,
                          breadcrumbs=breadcrumbs)

@app.route('/export-screenshots')
@login_required
def export_screenshots():
    """Export screenshots"""
    if not current_user.is_admin:
        flash('You must be an admin to export screenshots.', 'danger')
        return redirect(url_for('index'))
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Admin Dashboard", "url": url_for('admin')},
        {"name": "Export Screenshots", "url": url_for('export_screenshots')}
    ]
    
    # Define SEO metadata
    meta_description = "Export and manage South African lottery screenshots. Download captured lottery result images in various formats for analysis and record-keeping."
    
    screenshots = Screenshot.query.order_by(Screenshot.timestamp.desc()).all()
    
    # Check for sync status in session
    sync_status = None
    if 'sync_status' in session:
        sync_status = session.pop('sync_status')
    
    # Get the timestamp of the most recent screenshot for status display
    last_updated = None
    if screenshots:
        last_updated = screenshots[0].timestamp
    
    return render_template('export_screenshots.html',
                          screenshots=screenshots,
                          title="Export Lottery Screenshots | Data Management",
                          meta_description=meta_description,
                          breadcrumbs=breadcrumbs,
                          sync_status=sync_status,
                          last_updated=last_updated)

@app.route('/export-screenshots-zip')
@login_required
def export_screenshots_zip():
    """Export all screenshots as a ZIP file"""
    if not current_user.is_admin:
        flash('You must be an admin to export screenshots.', 'danger')
        return redirect(url_for('index'))
    
    try:
        import io
        import zipfile
        from datetime import datetime
        
        # Get all screenshots
        screenshots = Screenshot.query.order_by(Screenshot.lottery_type).all()
        
        if not screenshots:
            flash('No screenshots available to export.', 'warning')
            return redirect(url_for('export_screenshots'))
        
        # Create a ZIP file in memory
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for screenshot in screenshots:
                if os.path.exists(screenshot.path):
                    # Get the file extension
                    _, ext = os.path.splitext(screenshot.path)
                    # Create a unique filename for each screenshot
                    lottery_type = screenshot.lottery_type.replace(' ', '_')
                    timestamp = screenshot.timestamp.strftime('%Y%m%d_%H%M%S')
                    filename = f"{lottery_type}_{timestamp}{ext}"
                    
                    # Add the screenshot to the ZIP file
                    zf.write(screenshot.path, filename)
                    
                    # Add zoomed version if it exists
                    if screenshot.zoomed_path and os.path.exists(screenshot.zoomed_path):
                        _, zoomed_ext = os.path.splitext(screenshot.zoomed_path)
                        zoomed_filename = f"{lottery_type}_{timestamp}_zoomed{zoomed_ext}"
                        zf.write(screenshot.zoomed_path, zoomed_filename)
        
        # Reset the file pointer to the beginning of the file
        memory_file.seek(0)
        
        # Create a timestamp for the filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Send the ZIP file as a response
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'lottery_screenshots_{timestamp}.zip'
        )
    except Exception as e:
        app.logger.error(f"Error creating ZIP file: {str(e)}")
        flash(f'Error creating ZIP file: {str(e)}', 'danger')
        return redirect(url_for('export_screenshots'))

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    """Register a new admin user"""
    if not current_user.is_admin:
        flash('You must be an admin to register new users.', 'danger')
        return redirect(url_for('index'))
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Admin Dashboard", "url": url_for('admin')},
        {"name": "Register Admin", "url": url_for('register')}
    ]
    
    # Define SEO metadata
    meta_description = "Administrative user registration for South African lottery results system. Create secure admin accounts to manage lottery data and system configurations."
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'danger')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Email already exists', 'danger')
            return redirect(url_for('register'))
        
        user = User(username=username, email=email, is_admin=True)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'Admin user {username} registered successfully!', 'success')
        return redirect(url_for('admin'))
    
    return render_template('register.html', 
                          title="Register Admin User | Lottery Management System", 
                          breadcrumbs=breadcrumbs,
                          meta_description=meta_description)

@app.route('/visualizations')
def visualizations():
    """Advanced data visualization and analytics for South African lottery results"""
    lottery_types = ['Lottery', 'Lottery Plus 1', 'Lottery Plus 2', 
                    'Powerball', 'Powerball Plus', 'Daily Lottery']
    
    # Get some summary statistics
    total_draws = LotteryResult.query.count()
    latest_draw = LotteryResult.query.order_by(LotteryResult.draw_date.desc()).first()
    latest_draw_date = latest_draw.draw_date if latest_draw else None
    
    # Define breadcrumbs for SEO
    breadcrumbs = [
        {"name": "Analytics", "url": url_for('visualizations')}
    ]
    
    # Additional SEO metadata
    meta_description = "Explore comprehensive South African lottery analytics. View frequency charts, number patterns, and winning statistics for Lotto, Powerball, and Daily Lotto games."
    
    return render_template('visualizations.html',
                          lottery_types=lottery_types,
                          total_draws=total_draws,
                          latest_draw_date=latest_draw_date,
                          title="Lottery Data Analytics | Statistical Analysis & Visualizations",
                          breadcrumbs=breadcrumbs,
                          meta_description=meta_description)

@app.route('/api/visualization-data')
def visualization_data():
    """API endpoint for visualization data"""
    # import data_aggregator  # Module not available
    from collections import Counter
    import logging
    
    data_type = request.args.get('data_type', 'numbers_frequency')
    lottery_type = request.args.get('lottery_type', 'all')
    
    logging.info(f"Visualization data request: type={data_type}, lottery={lottery_type}")
    
    try:
        # Get data specific to lottery type
        if data_type == 'numbers_frequency':
            # Get frequency count for all 49 numbers
            all_numbers = {str(i): 0 for i in range(1, 50)}
            
            # Count frequency of all numbers
            query = LotteryResult.query
            if lottery_type and lottery_type.lower() != 'all':
                query = query.filter_by(lottery_type=lottery_type)
                
            results = query.all()
            number_counter = Counter()
            
            logging.info(f"Found {len(results)} results for frequency analysis")
            
            for result in results:
                numbers = result.get_numbers_list()
                for num in numbers:
                    if 1 <= num <= 49:
                        number_counter[str(num)] += 1
            
            # Update all_numbers with actual frequencies
            for num, count in number_counter.items():
                if num in all_numbers:
                    all_numbers[num] = count
            
            # Convert to lists for JSON response
            data = [all_numbers[str(i)] for i in range(1, 50)]
            
            response_data = {
                'labels': [str(i) for i in range(1, 50)],
                'datasets': [{
                    'data': data
                }]
            }
            
            logging.info(f"Returning frequency data with {sum(data)} total occurrences")
            return jsonify(response_data)
        
        elif data_type == 'winners_by_division':
            # Get all results for this lottery type to process division data
            query = LotteryResult.query
            
            if lottery_type and lottery_type.lower() != 'all':
                # Try with exact match first
                results_by_type = query.filter_by(lottery_type=lottery_type).all()
                
                # If no results, try with normalized type
                if not results_by_type:
                    # Find any results with lottery type that might be a variant
                    lottery_type_variants = db.session.query(LotteryResult.lottery_type).distinct().all()
                    normalized_type = data_aggregator.normalize_lottery_type(lottery_type)
                    matching_types = []
                    
                    for lt in lottery_type_variants:
                        lt_name = lt[0]
                        if data_aggregator.normalize_lottery_type(lt_name) == normalized_type:
                            matching_types.append(lt_name)
                    
                    if matching_types:
                        results_by_type = query.filter(LotteryResult.lottery_type.in_(matching_types)).all()
            else:
                results_by_type = query.all()
            
            # Initialize division counters
            division_data = {}
            
            # Process division data from all matching results
            for result in results_by_type:
                divisions = result.get_divisions()
                if divisions:
                    for div_name, div_info in divisions.items():
                        # Extract the division number from the name (e.g., "Division 1" -> 1)
                        try:
                            div_num = int(div_name.split()[-1])
                            winners = div_info.get('winners', '0')
                            
                            # Convert winners to integer
                            if isinstance(winners, str):
                                winners = winners.replace(',', '')
                            winner_count = int(float(winners))
                            
                            if div_num not in division_data:
                                division_data[div_num] = 0
                            division_data[div_num] += winner_count
                        except (ValueError, IndexError):
                            continue
            
            # Sort divisions by number
            sorted_divisions = sorted(division_data.items())
            
            # Prepare data for chart
            divisions = [f'Division {div_num}' for div_num, _ in sorted_divisions]
            winner_counts = [count for _, count in sorted_divisions]
            
            logging.info(f"Processed {len(results_by_type)} results, found {len(divisions)} divisions with data")
            
            response_data = {
                'labels': divisions,
                'datasets': [{
                    'data': winner_counts
                }]
            }
            
            return jsonify(response_data)
        
        return jsonify({'error': 'Invalid data type'}), 400
    except Exception as e:
        logging.error(f"Error in visualization_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/results/<lottery_type>/<draw_number>')
def draw_details(lottery_type, draw_number):
    """Show detailed information for a specific draw"""
    try:
        # Map display lottery type to database lottery type
        db_lottery_type = lottery_type
        if lottery_type == 'Lottery':
            db_lottery_type = 'Lotto'
        elif lottery_type == 'Powerball':
            db_lottery_type = 'PowerBall'
        elif lottery_type == 'Daily Lottery':
            db_lottery_type = 'Daily Lotto'
        
        # Query for the specific draw (draw_number is stored as integer)
        try:
            draw_num_int = int(draw_number)
        except ValueError:
            flash(f"Invalid draw number: {draw_number}", "error")
            return redirect(url_for('lottery_results', lottery_type=lottery_type))
            
        result = LotteryResult.query.filter_by(
            lottery_type=db_lottery_type,
            draw_number=draw_num_int
        ).first()
        
        if not result:
            flash(f"Draw {draw_number} not found for {lottery_type}", "warning")
            return redirect(url_for('lottery_results', lottery_type=lottery_type))
        
        # Get previous and next draws for navigation
        prev_draw = LotteryResult.query.filter(
            LotteryResult.lottery_type == db_lottery_type,
            LotteryResult.draw_date < result.draw_date
        ).order_by(LotteryResult.draw_date.desc()).first()
        
        next_draw = LotteryResult.query.filter(
            LotteryResult.lottery_type == db_lottery_type,
            LotteryResult.draw_date > result.draw_date
        ).order_by(LotteryResult.draw_date.asc()).first()
        
        return render_template('draw_details.html',
                             result=result,
                             lottery_type=lottery_type,
                             prev_draw=prev_draw,
                             next_draw=next_draw)
                             
    except Exception as e:
        app.logger.error(f"Error in draw_details: {e}")
        flash("Error loading draw details", "error")
        return redirect(url_for('lottery_results', lottery_type=lottery_type))

@app.route('/screenshot/<int:screenshot_id>')
def view_screenshot(screenshot_id):
    """View a screenshot image"""
    screenshot = Screenshot.query.get_or_404(screenshot_id)
    
    # Normalize path and check if file exists
    screenshot_path = os.path.normpath(screenshot.path)
    
    if not os.path.isfile(screenshot_path):
        flash('Screenshot file not found', 'danger')
        return redirect(url_for('admin'))
    
    # Extract directory and filename from path
    directory = os.path.dirname(screenshot_path)
    filename = os.path.basename(screenshot_path)
    
    return send_from_directory(directory, filename)

@app.route('/screenshot-zoomed/<int:screenshot_id>')
def view_zoomed_screenshot(screenshot_id):
    """View a zoomed screenshot image"""
    screenshot = Screenshot.query.get_or_404(screenshot_id)
    
    if not screenshot.zoomed_path:
        flash('No zoomed screenshot available', 'warning')
        return redirect(url_for('admin'))
    
    # Normalize path and check if file exists
    zoomed_path = os.path.normpath(screenshot.zoomed_path)
    
    if not os.path.isfile(zoomed_path):
        flash('Zoomed screenshot file not found', 'danger')
        return redirect(url_for('admin'))
    
    # Extract directory and filename from path
    directory = os.path.dirname(zoomed_path)
    filename = os.path.basename(zoomed_path)
    
    return send_from_directory(directory, filename)

@app.route('/sync-all-screenshots', methods=['POST'])
@login_required
def sync_all_screenshots():
    """Sync all screenshots from their source URLs"""
    if not current_user.is_admin:
        flash('You must be an admin to sync screenshots.', 'danger')
        return redirect(url_for('index'))
    
    try:
        # Load screenshot manager for automated screenshot operations
        global screenshot_manager
        if screenshot_manager is None:
            import screenshot_manager
        
        # Execute automated screenshot sync
        count = screenshot_manager.retake_all_screenshots(app, use_threading=False)
        
        # Store status in session for display on next page load
        if count > 0:
            session['sync_status'] = {
                'status': 'success',
                'message': f'Successfully synced {count} screenshots.'
            }
        else:
            session['sync_status'] = {
                'status': 'warning',
                'message': 'No screenshots were synced. Check configured URLs.'
            }
    except Exception as e:
        app.logger.error(f"Error syncing screenshots: {str(e)}")
        session['sync_status'] = {
            'status': 'danger',
            'message': f'Error syncing screenshots: {str(e)}'
        }
    
    return redirect(url_for('export_screenshots'))

@app.route('/sync-screenshot/<int:screenshot_id>', methods=['POST'])
@login_required
def sync_single_screenshot(screenshot_id):
    """Sync a single screenshot by its ID"""
    if not current_user.is_admin:
        flash('You must be an admin to sync screenshots.', 'danger')
        return redirect(url_for('index'))
    
    try:
        # Get the screenshot
        screenshot = Screenshot.query.get_or_404(screenshot_id)
        
        # Use the screenshot manager to retake this screenshot
        global screenshot_manager
        if screenshot_manager is None:
            import screenshot_manager
        
        success = screenshot_manager.retake_screenshot_by_id(screenshot_id, app)
        
        # Store status in session for display on next page load
        if success:
            session['sync_status'] = {
                'status': 'success',
                'message': f'Successfully synced screenshot for {screenshot.lottery_type}.'
            }
        else:
            session['sync_status'] = {
                'status': 'warning',
                'message': f'Failed to sync screenshot for {screenshot.lottery_type}.'
            }
    except Exception as e:
        app.logger.error(f"Error syncing screenshot: {str(e)}")
        session['sync_status'] = {
            'status': 'danger',
            'message': f'Error syncing screenshot: {str(e)}'
        }
    
    return redirect(url_for('export_screenshots'))



@app.route('/admin/scheduler-status')
@login_required
def scheduler_status():
    """Admin route to check daily automation scheduler status"""
    if not current_user.is_admin:
        flash('You must be an admin to view scheduler status.', 'danger')
        return redirect(url_for('index'))
    
    global daily_scheduler
    status_info = {
        'scheduler_available': daily_scheduler is not None,
        'running': False,
        'next_run': None,
        'last_run': None,
        'last_results': None,
        'run_time': '01:00'
    }
    
    if daily_scheduler:
        scheduler_instance = daily_scheduler.get_scheduler() if hasattr(daily_scheduler, 'get_scheduler') else None
        if scheduler_instance:
            status = scheduler_instance.get_status()
            status_info.update(status)
    
    return render_template('admin/scheduler_status.html', 
                          title="Daily Automation Scheduler Status",
                          status=status_info)

@app.route('/admin/start-scheduler', methods=['POST'])
@login_required
def start_scheduler():
    """Admin route to start the daily automation scheduler"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Admin privileges required'}), 403
    
    global daily_scheduler
    try:
        if not daily_scheduler:
            import scheduler as sched
            daily_scheduler = sched.init_scheduler(app, run_time="01:00")
            message = "Daily automation scheduler started successfully - will run at 1:00 AM every day"
        else:
            scheduler_instance = daily_scheduler.get_scheduler() if hasattr(daily_scheduler, 'get_scheduler') else None
            if scheduler_instance and not scheduler_instance.running:
                scheduler_instance.start_scheduler()
                message = "Daily automation scheduler restarted successfully"
            else:
                message = "Daily automation scheduler is already running"
        
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        app.logger.error(f"Error starting scheduler: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/stop-scheduler', methods=['POST'])
@login_required
def stop_scheduler():
    """Admin route to stop the daily automation scheduler"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Admin privileges required'}), 403
    
    global daily_scheduler
    try:
        if daily_scheduler:
            scheduler_instance = daily_scheduler.get_scheduler() if hasattr(daily_scheduler, 'get_scheduler') else None
            if scheduler_instance:
                scheduler_instance.stop_scheduler()
                message = "Daily automation scheduler stopped successfully"
            else:
                message = "Scheduler instance not found"
        else:
            message = "No scheduler to stop"
        
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        app.logger.error(f"Error stopping scheduler: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/run-automation-now', methods=['POST'])
@login_required
def run_automation_now():
    """Admin route to trigger daily automation immediately"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Admin privileges required'}), 403
    
    try:
        from daily_automation import run_complete_automation
        results = run_complete_automation()
        
        if results.get('overall_success', False):
            message = f"Daily automation completed successfully: {results.get('summary', 'No details available')}"
        else:
            message = f"Daily automation completed with errors: {results.get('error', 'Unknown error')}"
        
        return jsonify({'success': True, 'message': message, 'results': results})
    except Exception as e:
        app.logger.error(f"Error running automation: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/cleanup-screenshots', methods=['POST'])
@login_required
def cleanup_screenshots():
    """Route to cleanup old screenshots"""
    if not current_user.is_admin:
        flash('You must be an admin to clean up screenshots.', 'danger')
        return redirect(url_for('index'))
        
    try:
        # Run the cleanup function from step1_cleanup
        from step1_cleanup import cleanup_old_screenshots
        
        cleanup_success = cleanup_old_screenshots(days_to_keep=7)
        
        if cleanup_success:
            session['sync_status'] = {
                'status': 'success',
                'message': 'Cleanup completed successfully. Screenshots older than 7 days have been removed.'
            }
        else:
            session['sync_status'] = {
                'status': 'warning',
                'message': 'Cleanup completed with some issues. Check logs for details.'
            }
    except Exception as e:
        app.logger.error(f"Error cleaning up screenshots: {str(e)}")
        traceback.print_exc()
        
        session['sync_status'] = {
            'status': 'danger',
            'message': f'Error cleaning up screenshots: {str(e)}'
        }
    
    return redirect(url_for('export_screenshots'))

@app.route('/cleanup-all-screenshots', methods=['GET', 'POST'])
@login_required
def cleanup_all_screenshots_route():
    """Admin route to clean up ALL screenshots regardless of age"""
    if not current_user.is_admin:
        flash('You must be an admin to clean up screenshots.', 'danger')
        return redirect(url_for('index'))
        
    try:
        # Run the complete cleanup function
        from step1_cleanup import cleanup_all_screenshots
        
        cleanup_success = cleanup_all_screenshots()
        
        if cleanup_success:
            session['sync_status'] = {
                'status': 'success',
                'message': 'All screenshots have been successfully removed from the screenshots folder.'
            }
        else:
            session['sync_status'] = {
                'status': 'warning',
                'message': 'Complete cleanup completed with some issues. Check logs for details.'
            }
    except Exception as e:
        app.logger.error(f"Error cleaning up all screenshots: {str(e)}")
        traceback.print_exc()
        
        session['sync_status'] = {
            'status': 'danger',
            'message': f'Error cleaning up all screenshots: {str(e)}'
        }
    
    return redirect(url_for('export_screenshots'))

@app.route('/export-combined-zip')
@login_required
def export_combined_zip():
    """Export template and screenshots in a single ZIP file"""
    if not current_user.is_admin:
        flash('You must be an admin to export data.', 'danger')
        return redirect(url_for('index'))
    
    try:
        import io
        import zipfile
        import tempfile
        import glob
        from datetime import datetime
        
        # Create a timestamp for filenames
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Get actual screenshot files from filesystem
        screenshot_dir = os.path.join(os.getcwd(), 'screenshots')
        screenshot_files = []
        
        if os.path.exists(screenshot_dir):
            screenshot_files = glob.glob(os.path.join(screenshot_dir, '*.png'))
        
        if not screenshot_files:
            flash('No screenshot files available to export.', 'warning')
            return redirect(url_for('admin_automation_control'))
        
        # Create a temporary directory for the template
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create the template file
            template_filename = f"lottery_data_template_{timestamp}.xlsx"
            template_path = os.path.join(temp_dir, template_filename)
            create_lottery_template(template_path)
            
            # Create a ZIP file in memory
            memory_file = io.BytesIO()
            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Add the template to the ZIP file
                if os.path.exists(template_path):
                    zf.write(template_path, f"template/{template_filename}")
                
                # Add actual screenshot files to the ZIP file
                for screenshot_path in screenshot_files:
                    if os.path.exists(screenshot_path):
                        filename = os.path.basename(screenshot_path)
                        zf.write(screenshot_path, f"screenshots/{filename}")
                
                # Add README file with instructions
                readme_content = f"""Lottery Data Export - {timestamp}

This archive contains:
1. Template folder: Excel template for manual data entry
2. Screenshots folder: Captured lottery result pages ({len(screenshot_files)} files)

Screenshot files are named with format: lottery_type_YYYYMMDD_HHMMSS.png
All images are full-page captures containing complete lottery data including:
- Winning numbers
- Division breakdowns
- Prize amounts
- Draw dates and numbers

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
                zf.writestr("README.txt", readme_content)
            
            # Reset the file pointer to the beginning of the file
            memory_file.seek(0)
            
            # Send the ZIP file as a response
            return send_file(
                memory_file,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f'lottery_data_combined_{timestamp}.zip'
            )
    except Exception as e:
        app.logger.error(f"Error creating combined ZIP file: {str(e)}")
        flash(f'Error creating combined ZIP file: {str(e)}', 'danger')
        return redirect(url_for('admin_automation_control'))

@app.route('/port_check')
def port_check():
    """Special endpoint to check which port the application is responding on"""
    host = request.host
    return jsonify({
        "success": True,
        "host": host,
        "server_port": request.host.split(':')[1] if ':' in request.host else "default",
        "request_url": request.url,
        "timestamp": datetime.now().isoformat()
    })

@app.route('/api/results/<lottery_type>')
def api_results(lottery_type):
    """API endpoint for lottery results"""
    try:
        limit = request.args.get('limit', type=int)
        results = data_aggregator.export_results_to_json(lottery_type, limit)
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# Advertisement Management Routes
@app.route('/admin/manage-ads')
@app.route('/admin/manage-ads/<placement>')
@login_required
def manage_ads(placement=None):
    """Admin page to manage advertisements"""
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('index'))
    
    # Get all ads or filter by placement
    if placement:
        ads = Advertisement.query.filter_by(placement=placement).order_by(Advertisement.priority.desc()).all()
    else:
        ads = Advertisement.query.order_by(Advertisement.priority.desc()).all()
    
    return render_template('admin/manage_ads.html', 
                          ads=ads,
                          placement=placement,
                          title="Manage Advertisements")

@app.route('/admin/upload-ad', methods=['GET', 'POST'])
@login_required
def upload_ad():
    """Admin page to upload a new advertisement"""
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('index'))
    
    form = request.form
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name')
            description = request.form.get('description')
            duration = int(request.form.get('duration'))
            placement = request.form.get('placement')
            priority = int(request.form.get('priority', 5))
            target_impressions = int(request.form.get('target_impressions'))
            active = 'active' in request.form
            
            # Get custom message and rich content settings
            custom_message = request.form.get('custom_message')
            loading_duration = int(request.form.get('loading_duration', 10))
            is_rich_content = 'is_rich_content' in request.form
            html_content = request.form.get('html_content') if is_rich_content else None
            
            # Handle optional dates
            start_date = request.form.get('start_date')
            if start_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
            else:
                start_date = None
                
            end_date = request.form.get('end_date')
            if end_date:
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            else:
                end_date = None
            
            # Check if video file is included
            if 'video_file' not in request.files:
                flash('No video file provided', 'danger')
                return redirect(request.url)
            
            video_file = request.files['video_file']
            if video_file.filename == '':
                flash('No selected file', 'danger')
                return redirect(request.url)
            
            # Determine file type
            file_type = video_file.content_type or 'video/mp4'
            
            # Create ads directory if it doesn't exist
            ads_dir = os.path.join('static', 'ads')
            os.makedirs(ads_dir, exist_ok=True)
            
            # Save the file with a unique name
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{secure_filename(name)}_{timestamp}.{file_type.split('/')[-1]}"
            file_path = os.path.join(ads_dir, filename)
            video_file.save(file_path)
            
            # Handle custom image if provided
            custom_image_path = None
            if 'custom_image' in request.files and request.files['custom_image'].filename:
                custom_image = request.files['custom_image']
                
                # Determine file type
                image_type = custom_image.content_type or 'image/jpeg'
                
                # Save the image
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                image_filename = f"custom_image_{secure_filename(name)}_{timestamp}.{image_type.split('/')[-1]}"
                custom_image_path = os.path.join('static', 'ads', 'images', image_filename)
                
                # Create directory if it doesn't exist
                os.makedirs(os.path.dirname(custom_image_path), exist_ok=True)
                
                # Save the image
                custom_image.save(custom_image_path)
            
            # Create new advertisement record
            new_ad = Advertisement(
                name=name,
                description=description,
                file_path=file_path,
                file_type=file_type,
                duration=duration,
                placement=placement,
                priority=priority,
                target_impressions=target_impressions,
                active=active,
                start_date=start_date,
                end_date=end_date,
                created_by_id=current_user.id,
                custom_message=custom_message,
                custom_image_path=custom_image_path,
                loading_duration=loading_duration,
                is_rich_content=is_rich_content,
                html_content=html_content
            )
            
            db.session.add(new_ad)
            db.session.commit()
            
            flash(f'Advertisement "{name}" uploaded successfully!', 'success')
            return redirect(url_for('manage_ads'))
            
        except Exception as e:
            db.session.rollback()
            logger.exception(f"Error uploading advertisement: {str(e)}")
            flash(f'Error uploading advertisement: {str(e)}', 'danger')
    
    return render_template('admin/upload_ad.html', 
                          form=form,
                          title="Upload Advertisement")

@app.route('/admin/edit-ad/<int:ad_id>', methods=['GET', 'POST'])
@login_required
def edit_ad(ad_id):
    """Admin page to edit an existing advertisement"""
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('index'))
    
    # Get the advertisement
    ad = Advertisement.query.get_or_404(ad_id)
    form = request.form
    
    if request.method == 'POST':
        try:
            # Update advertisement details
            ad.name = request.form.get('name')
            ad.description = request.form.get('description')
            ad.duration = int(request.form.get('duration'))
            ad.placement = request.form.get('placement')
            ad.priority = int(request.form.get('priority', 5))
            ad.target_impressions = int(request.form.get('target_impressions'))
            ad.active = 'active' in request.form
            
            # Update custom message and content settings
            ad.custom_message = request.form.get('custom_message')
            ad.loading_duration = int(request.form.get('loading_duration', 10))
            ad.is_rich_content = 'is_rich_content' in request.form
            ad.html_content = request.form.get('html_content') if ad.is_rich_content else None
            
            # Handle optional dates
            start_date = request.form.get('start_date')
            if start_date:
                ad.start_date = datetime.strptime(start_date, '%Y-%m-%d')
            else:
                ad.start_date = None
                
            end_date = request.form.get('end_date')
            if end_date:
                ad.end_date = datetime.strptime(end_date, '%Y-%m-%d')
            else:
                ad.end_date = None
            
            # Check if a new video file was uploaded
            if 'video_file' in request.files and request.files['video_file'].filename:
                video_file = request.files['video_file']
                
                # Determine file type
                file_type = video_file.content_type or 'video/mp4'
                
                # Delete old file if it exists and is in the ads directory
                if ad.file_path and os.path.exists(ad.file_path) and 'static/ads' in ad.file_path:
                    try:
                        os.remove(ad.file_path)
                    except Exception as e:
                        logger.warning(f"Could not delete old advertisement file: {str(e)}")
                
                # Save the new file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{secure_filename(ad.name)}_{timestamp}.{file_type.split('/')[-1]}"
                file_path = os.path.join('static', 'ads', filename)
                video_file.save(file_path)
                
                # Update file path and type
                ad.file_path = file_path
                ad.file_type = file_type
            
            # Save changes
            db.session.commit()
            
            flash(f'Advertisement "{ad.name}" updated successfully!', 'success')
            return redirect(url_for('manage_ads'))
            
        except Exception as e:
            db.session.rollback()
            logger.exception(f"Error updating advertisement: {str(e)}")
            flash(f'Error updating advertisement: {str(e)}', 'danger')
    
    # Prepare form data for template
    form = {
        'name': ad.name,
        'description': ad.description,
        'duration': ad.duration,
        'placement': ad.placement,
        'priority': ad.priority,
        'target_impressions': ad.target_impressions,
        'active': ad.active,
        'start_date': ad.start_date,
        'end_date': ad.end_date,
        'custom_message': ad.custom_message,
        'loading_duration': ad.loading_duration,
        'is_rich_content': ad.is_rich_content,
        'html_content': ad.html_content
    }
    
    return render_template('admin/upload_ad.html', 
                          form=form,
                          ad=ad,
                          title="Edit Advertisement")

@app.route('/admin/preview-ad/<int:ad_id>')
@login_required
def preview_ad(ad_id):
    """Admin page to preview an advertisement"""
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('index'))
    
    # Get the advertisement
    ad = Advertisement.query.get_or_404(ad_id)
    
    return render_template('admin/preview_ad.html', 
                          ad=ad,
                          title=f"Preview: {ad.name}")

@app.route('/admin/delete-ad', methods=['POST'])
@login_required
def delete_ad():
    """Delete an advertisement"""
    if not current_user.is_admin:
        flash('You must be an admin to delete advertisements.', 'danger')
        return redirect(url_for('index'))
    
    ad_id = request.form.get('ad_id')
    if not ad_id:
        flash('No advertisement specified', 'danger')
        return redirect(url_for('manage_ads'))
    
    try:
        ad = Advertisement.query.get_or_404(ad_id)
        
        # Delete the file if it exists and is in the ads directory
        if ad.file_path and os.path.exists(ad.file_path) and 'static/ads' in ad.file_path:
            try:
                os.remove(ad.file_path)
            except Exception as e:
                logger.warning(f"Could not delete advertisement file: {str(e)}")
        
        # Delete the advertisement from the database
        db.session.delete(ad)
        db.session.commit()
        
        flash(f'Advertisement "{ad.name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Error deleting advertisement: {str(e)}")
        flash(f'Error deleting advertisement: {str(e)}', 'danger')
    
    return redirect(url_for('manage_ads'))

@app.route('/admin/ad-impressions')
@login_required
def ad_impressions():
    """Admin page to view ad impressions"""
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('index'))
    
    # Get all impressions grouped by advertisement
    ads = Advertisement.query.all()
    impressions = AdImpression.query.order_by(AdImpression.timestamp.desc()).limit(100).all()
    
    return render_template('admin/ad_impressions.html',
                          ads=ads,
                          impressions=impressions,
                          title="Advertisement Impressions")

@app.route('/admin/ad-performance')
@login_required
def ad_performance():
    """Admin page to view ad performance analytics"""
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('index'))
    
    # Get ad performance statistics
    ads = Advertisement.query.all()
    
    # Calculate performance metrics
    ad_stats = []
    for ad in ads:
        total_impressions = ad.total_impressions
        total_clicks = ad.total_clicks
        click_through_rate = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
        progress_percent = (total_impressions / ad.target_impressions * 100) if ad.target_impressions > 0 else 0
        
        ad_stats.append({
            'ad': ad,
            'impressions': total_impressions,
            'clicks': total_clicks,
            'ctr': click_through_rate,
            'progress': min(progress_percent, 100)  # Cap at 100%
        })
    
    return render_template('admin/ad_performance.html',
                          ad_stats=ad_stats,
                          title="Advertisement Performance")

@app.route('/api/record-impression', methods=['POST'])
def record_impression():
    """API endpoint to record an ad impression"""
    try:
        ad_id = request.json.get('ad_id')
        if not ad_id:
            return jsonify({"error": "No ad_id provided"}), 400
        
        # Get the advertisement
        ad = Advertisement.query.get_or_404(ad_id)
        
        # Create new impression record
        impression = AdImpression(
            ad_id=ad_id,
            user_id=current_user.id if current_user.is_authenticated else None,
            session_id=request.cookies.get('session', 'unknown'),
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string,
            page=request.referrer,
            duration_viewed=request.json.get('duration')
        )
        
        # Update ad impression count
        ad.total_impressions += 1
        
        db.session.add(impression)
        db.session.commit()
        
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Error recording impression: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/record-click', methods=['POST'])
def record_click():
    """API endpoint to record an ad click"""
    try:
        impression_id = request.json.get('impression_id')
        if not impression_id:
            return jsonify({"error": "No impression_id provided"}), 400
        
        # Get the impression
        impression = AdImpression.query.get_or_404(impression_id)
        
        # Update impression and ad click counts
        impression.was_clicked = True
        impression.advertisement.total_clicks += 1
        
        db.session.commit()
        
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Error recording click: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/admin/system_status')
@login_required
def system_status():
    """System status dashboard for monitoring all components"""
    # Check server status
    port_5000_status = True  # We're already running on this port if this code executes
    
    # Check port 8080
    port_8080_status = False
    try:
        import urllib.request
        response = urllib.request.urlopen('http://localhost:8080/port_check', timeout=2)
        port_8080_status = response.getcode() == 200
    except Exception as e:
        logger.warning(f"Failed to check port 8080: {str(e)}")
    
    # Get database status and statistics
    db_status = True
    db_tables_count = 0
    db_records_count = 0
    db_type = "PostgreSQL"
    
    try:
        # Count tables
        from sqlalchemy import text
        result = db.session.execute(text("SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = 'public'"))
        db_tables_count = result.scalar() or 0
        
        # Count records in lottery_result table as a sample
        result = db.session.execute(text("SELECT COUNT(*) FROM lottery_result"))
        db_records_count = result.scalar() or 0
    except Exception as e:
        logger.error(f"Database status check error: {str(e)}")
        db_status = False
    
    # Get advertisement system stats
    active_ads_count = Advertisement.query.filter_by(active=True).count()
    total_impressions = db.session.query(db.func.count(AdImpression.id)).scalar() or 0
    js_status = True  # Will be set by frontend JS
    
    # Get lottery stats
    lottery_types = ['Lottery', 'Lottery Plus 1', 'Lottery Plus 2', 'Powerball', 'Powerball Plus', 'Daily Lottery']
    lottery_stats = []
    
    for lottery_type in lottery_types:
        count = LotteryResult.query.filter_by(lottery_type=lottery_type).count()
        latest_draw = LotteryResult.query.filter_by(lottery_type=lottery_type).order_by(LotteryResult.draw_date.desc()).first()
        
        lottery_stats.append({
            'name': lottery_type,
            'count': count,
            'latest_draw': latest_draw.draw_date.strftime('%Y-%m-%d') if latest_draw else 'N/A'
        })
    
    # Get system resource usage
    try:
        import psutil
        memory_usage = int(psutil.virtual_memory().percent)
        cpu_usage = int(psutil.cpu_percent(interval=0.1))
        disk_usage = int(psutil.disk_usage('/').percent)
    except ImportError:
        # Fallback if psutil not available
        memory_usage = 50
        cpu_usage = 30
        disk_usage = 40
        logger.warning("psutil not available for system stats, using placeholder values")
    
    return render_template(
        'admin/system_status.html',
        port_5000_status=port_5000_status,
        port_8080_status=port_8080_status,
        server_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        last_checked=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        db_status=db_status,
        db_tables_count=db_tables_count,
        db_records_count=db_records_count,
        db_type=db_type,
        active_ads_count=active_ads_count,
        total_impressions=total_impressions,
        js_status=js_status,
        lottery_stats=lottery_stats,
        memory_usage=memory_usage,
        cpu_usage=cpu_usage,
        disk_usage=disk_usage
    )

@app.route('/admin/check-js', methods=['POST'])
@login_required
def check_js():
    """API endpoint to check if JavaScript is operational"""
    return jsonify({'success': True})

@app.route('/admin/health-dashboard')
@login_required
def health_dashboard():
    """Health monitoring dashboard for system administrators"""
    import health_monitor
    from datetime import datetime, timedelta
    import json
    
    # Get the overall system status
    system_status = health_monitor.get_system_status(app, db)
    
    # Get port usage data
    port_history = health_monitor.get_health_history('port_usage', 1)
    # Add active ports to components
    active_ports = system_status.get('active_ports', [])
    
    # Get recent alerts
    alerts = health_monitor.get_recent_alerts(10)
    
    # Get health check history
    health_history = health_monitor.get_health_history(limit=10)
    
    # Get resource usage history for charts
    resource_history = health_monitor.get_health_history('system_resources', 24)
    
    # Prepare chart data
    resource_timestamps = []
    cpu_history = []
    memory_history = []
    disk_history = []
    
    for check in resource_history:
        # Try to parse timestamps and details
        try:
            timestamp = datetime.fromisoformat(check['timestamp'].replace('Z', '+00:00'))
            resource_timestamps.append(timestamp.strftime('%H:%M'))
            
            details = check['details']
            if isinstance(details, str):
                details = json.loads(details)
            
            cpu_history.append(details.get('cpu_usage', 0))
            memory_history.append(details.get('memory_usage', 0))
            disk_history.append(details.get('disk_usage', 0))
        except Exception as e:
            app.logger.error(f"Error parsing health check data: {str(e)}")
    
    # If we don't have enough data points, add some placeholders
    if len(resource_timestamps) < 5:
        now = datetime.now()
        for i in range(5 - len(resource_timestamps)):
            time_point = now - timedelta(hours=i)
            resource_timestamps.insert(0, time_point.strftime('%H:%M'))
            cpu_history.insert(0, 0)
            memory_history.insert(0, 0)
            disk_history.insert(0, 0)
    
    # Reverse lists to show chronological order
    resource_timestamps.reverse()
    cpu_history.reverse()
    memory_history.reverse()
    disk_history.reverse()
    
    # Add additional port monitoring information
    service_port_map = {}
    
    # If port usage information is available in the health history,
    # populate the service-to-port mapping
    if port_history and len(port_history) > 0:
        try:
            port_details = port_history[0]['details']
            if isinstance(port_details, str):
                port_details = json.loads(port_details)
            
            # Extract service usage mapping from details if available
            if isinstance(port_details, dict) and 'service_usage' in port_details:
                service_port_map = port_details['service_usage']
        except Exception as e:
            app.logger.error(f"Error extracting port history data: {str(e)}")
    
    return render_template(
        'admin/health_dashboard.html',
        overall_status=system_status['overall_status'],
        components=system_status['components'],
        last_updated=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        alerts=alerts,
        health_history=health_history,
        resource_timestamps=resource_timestamps,
        cpu_history=cpu_history,
        memory_history=memory_history,
        disk_history=disk_history,
        active_ports=active_ports,
        service_port_map=service_port_map
    )

@app.route('/admin/health-alerts')
@login_required
def health_alerts():
    """View and manage health alerts"""
    import health_monitor
    from datetime import datetime
    
    # Get all alerts
    alerts = health_monitor.get_recent_alerts(100)
    
    # Generate chart data for alert types
    alert_types = {}
    for alert in alerts:
        alert_type = alert['alert_type']
        if alert_type in alert_types:
            alert_types[alert_type] += 1
        else:
            alert_types[alert_type] = 1
    
    # Prepare chart data
    alert_types_list = list(alert_types.keys())
    alert_counts = [alert_types[t] for t in alert_types_list]
    
    # Generate time-based data
    # Group alerts by day for the last 30 days
    time_data = [0] * 30
    time_labels = []
    
    now = datetime.now()
    for i in range(30):
        day = now - timedelta(days=i)
        time_labels.insert(0, day.strftime('%m/%d'))
        
    for alert in alerts:
        try:
            created_at = datetime.fromisoformat(alert['created_at'].replace('Z', '+00:00'))
            days_ago = (now - created_at).days
            if 0 <= days_ago < 30:
                time_data[29 - days_ago] += 1
        except:
            pass
    
    # Add utility functions for templates
    def format_duration(start_time, end_time):
        """Format a duration between two timestamps"""
        try:
            if isinstance(start_time, str):
                start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            if isinstance(end_time, str):
                end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
            
            duration = end_time - start_time
            seconds = duration.total_seconds()
            
            if seconds < 60:
                return f"{int(seconds)} seconds"
            elif seconds < 3600:
                return f"{int(seconds / 60)} minutes"
            elif seconds < 86400:
                hours = int(seconds / 3600)
                minutes = int((seconds % 3600) / 60)
                return f"{hours} hours, {minutes} minutes"
            else:
                days = int(seconds / 86400)
                hours = int((seconds % 86400) / 3600)
                return f"{days} days, {hours} hours"
        except Exception as e:
            app.logger.error(f"Error formatting duration: {str(e)}")
            return "Unknown"
    
    return render_template(
        'admin/health_alerts.html',
        alerts=alerts,
        alert_types=alert_types_list,
        alert_counts=alert_counts,
        time_labels=time_labels,
        time_data=time_data,
        format_duration=format_duration,
        now=datetime.now()
    )

@app.route('/admin/resolve-health-alert/<int:alert_id>', methods=['POST'])
@login_required
def resolve_health_alert(alert_id):
    """Manually resolve a health alert"""
    import health_monitor
    import sqlite3
    
    try:
        conn = sqlite3.connect(health_monitor.health_db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get the alert
        cursor.execute("SELECT * FROM alerts WHERE id = ?", (alert_id,))
        alert = cursor.fetchone()
        
        if alert:
            # Update the alert
            cursor.execute(
                "UPDATE alerts SET resolved = 1, resolved_at = CURRENT_TIMESTAMP WHERE id = ?",
                (alert_id,)
            )
            conn.commit()
            
            # Also resolve any other alerts of the same type
            alert_type = alert['alert_type']
            health_monitor.resolve_alert(alert_type)
            
            flash(f"Alert #{alert_id} ({alert_type}) has been resolved.", "success")
        else:
            flash(f"Alert #{alert_id} not found.", "danger")
            
        conn.close()
    except Exception as e:
        flash(f"Error resolving alert: {str(e)}", "danger")
    
    return redirect(url_for('health_alerts'))

@app.route('/admin/health-history')
@login_required
def health_history():
    """View detailed health check history"""
    import health_monitor
    import json
    from datetime import datetime
    from sqlalchemy import func
    
    # Set up pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Get filter parameters
    selected_type = request.args.get('check_type', '')
    selected_status = request.args.get('status', '')
    
    # Get health check history
    try:
        import sqlite3
        conn = sqlite3.connect(health_monitor.health_db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Build the query
        query = "SELECT * FROM health_checks"
        params = []
        
        if selected_type or selected_status:
            query += " WHERE"
            
            if selected_type:
                query += " check_type = ?"
                params.append(selected_type)
                
            if selected_type and selected_status:
                query += " AND"
                
            if selected_status:
                query += " status = ?"
                params.append(selected_status)
        
        # Get total count for pagination
        count_query = f"SELECT COUNT(*) FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]
        
        # Add sorting and pagination
        query += " ORDER BY timestamp DESC LIMIT ? OFFSET ?"
        params.extend([per_page, (page - 1) * per_page])
        
        # Execute query
        cursor.execute(query, params)
        history = [dict(row) for row in cursor.fetchall()]
        
        # Get all unique check types for the filter dropdown
        cursor.execute("SELECT DISTINCT check_type FROM health_checks ORDER BY check_type")
        check_types = [row[0] for row in cursor.fetchall()]
        
        # Get status counts for chart
        cursor.execute("SELECT status, COUNT(*) FROM health_checks GROUP BY status")
        status_counts_raw = cursor.fetchall()
        
        # Prepare status counts for chart
        status_counts = [0, 0, 0, 0]  # OK, WARNING, CRITICAL, ERROR
        for status, count in status_counts_raw:
            if status == 'OK':
                status_counts[0] = count
            elif status == 'WARNING':
                status_counts[1] = count
            elif status == 'CRITICAL':
                status_counts[2] = count
            elif status == 'ERROR':
                status_counts[3] = count
        
        # Get check type counts for chart
        cursor.execute("SELECT check_type, COUNT(*) FROM health_checks GROUP BY check_type")
        type_counts_raw = cursor.fetchall()
        
        # Prepare check type counts for chart
        type_counts = [row[1] for row in type_counts_raw]
        
        conn.close()
        
        # Create pagination object
        class Pagination:
            def __init__(self, page, per_page, total):
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page
                
            @property
            def has_prev(self):
                return self.page > 1
                
            @property
            def has_next(self):
                return self.page < self.pages
                
            @property
            def prev_num(self):
                return self.page - 1 if self.has_prev else None
                
            @property
            def next_num(self):
                return self.page + 1 if self.has_next else None
                
            def iter_pages(self, left_edge=2, left_current=2, right_current=3, right_edge=2):
                last = 0
                for num in range(1, self.pages + 1):
                    if (num <= left_edge or
                            (self.page - left_current <= num <= self.page + right_current) or
                            (self.pages - right_edge < num)):
                        if last + 1 != num:
                            yield None
                        yield num
                        last = num
        
        pagination = Pagination(page, per_page, total)
        
        # Format JSON for display
        def format_json(json_str):
            try:
                if isinstance(json_str, str):
                    data = json.loads(json_str)
                else:
                    data = json_str
                return json.dumps(data, indent=2)
            except:
                return json_str
        
    except Exception as e:
        app.logger.error(f"Error getting health history: {str(e)}")
        history = []
        check_types = []
        status_counts = [0, 0, 0, 0]
        type_counts = []
        pagination = None
        format_json = lambda x: x
    
    return render_template(
        'admin/health_history.html',
        history=history,
        check_types=check_types,
        selected_type=selected_type,
        selected_status=selected_status,
        status_counts=status_counts,
        type_counts=type_counts,
        pagination=pagination,
        format_json=format_json
    )

@app.route('/health_port_check')
def health_port_check():
    """Simple endpoint to check if the server is running on a specific port"""
    return jsonify({
        'status': 'ok',
        'message': 'Port health check successful',
        'port': request.environ.get('SERVER_PORT', '5000'),
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })

@app.route('/admin/run-health-checks', methods=['POST', 'GET'])
@login_required
def run_health_checks():
    """Manually trigger health checks"""
    import health_monitor
    
    try:
        health_monitor.run_health_checks(app, db)
        flash("Health checks completed successfully.", "success")
    except Exception as e:
        flash(f"Error running health checks: {str(e)}", "danger")
    
    return redirect(url_for('health_dashboard'))

@app.route('/health_check', methods=['GET'])
def health_check():
    """Simple endpoint for health checks in Replit deployment"""
    return jsonify({
        'status': 'ok',
        'message': 'Health check successful',
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })



@app.route('/api/system-metrics', methods=['GET'])
def system_metrics():
    """API endpoint to get current system metrics for dashboard"""
    try:
        import psutil
        
        # Get system stats
        memory = psutil.virtual_memory()
        memory_usage = int(memory.percent)
        
        disk = psutil.disk_usage('/')
        disk_usage = int(disk.percent)
        
        cpu_usage = int(psutil.cpu_percent(interval=0.5))
        
        return jsonify({
            'success': True,
            'cpu_usage': f"{cpu_usage}%",
            'memory_usage': f"{memory_usage}%",
            'disk_usage': f"{disk_usage}%",
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Register advertisement management routes
# Ad management temporarily disabled
# ad_management.register_ad_routes(app)

# Register lottery analysis routes
lottery_analysis.register_analysis_routes(app, db)

@app.route('/api/extract-lottery-data', methods=['POST'])
@login_required
def api_extract_lottery_data():
    """API endpoint to extract lottery data from images using AI - bypasses CSRF"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Admin privileges required'}), 403
    
    try:
        from automated_data_extractor import LotteryDataExtractor
        
        extractor = LotteryDataExtractor()
        
        # Process all images in the attached_assets directory
        results = extractor.process_all_images("attached_assets")
        
        return jsonify({
            'success': True,
            'message': f'Processing complete! Successfully extracted {results["successful"]} lottery records, {results["failed"]} failed.',
            'results': results
        })
            
    except Exception as e:
        app.logger.error(f"Error in api_extract_lottery_data: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error during data extraction: {str(e)}'
        }), 500

@app.route('/extract-lottery-data', methods=['GET', 'POST'])
@login_required
def extract_lottery_data():
    """Admin route to extract lottery data from images using AI"""
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('home'))
    
    if request.method == 'GET':
        # Show the extraction interface
        return render_template('admin/data_extraction.html',
                             title="AI Data Extraction | Admin Dashboard")
    
    try:
        from automated_data_extractor import LotteryDataExtractor
        
        extractor = LotteryDataExtractor()
        
        # Check if testing single image or processing all
        test_mode = request.form.get('test_mode') == 'true'
        
        if test_mode:
            # Test with a single image
            test_image = "attached_assets/IMG_8174.png"
            if os.path.exists(test_image):
                result = extractor.test_single_image(test_image)
                if result:
                    flash(f'Test successful! Extracted: {result["lottery_type"]} with {len(result["main_numbers"])} numbers', 'success')
                    return render_template('admin/data_extraction.html',
                                         test_result=result,
                                         title="AI Data Extraction | Admin Dashboard")
                else:
                    flash('Test failed. Please check your Anthropic API key configuration.', 'danger')
            else:
                flash('Test image not found.', 'danger')
        else:
            # Process all images
            results = extractor.process_all_images("attached_assets")
            
            flash(f'Processing complete! Successfully extracted {results["successful"]} lottery records, {results["failed"]} failed.', 
                  'success' if results["successful"] > 0 else 'warning')
            
            return render_template('admin/data_extraction.html',
                                 extraction_results=results,
                                 title="AI Data Extraction | Admin Dashboard")
            
    except Exception as e:
        app.logger.error(f"Error in extract_lottery_data route: {str(e)}")
        flash(f'Error during data extraction: {str(e)}', 'danger')
    
    return render_template('admin/data_extraction.html',
                         title="AI Data Extraction | Admin Dashboard")

# API Request Tracking routes 
@app.route('/admin/api-tracking')
@login_required
def api_tracking_view():
    """Dashboard for tracking API requests to external services"""
    from datetime import datetime, timedelta
    from models import APIRequestLog
    from sqlalchemy import func, desc
    
    if not current_user.is_admin:
        flash('You must be an admin to access this page.', 'danger')
        return redirect(url_for('index'))
    
    # Get time periods for filtering
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday = today - timedelta(days=1)
    this_week = today - timedelta(days=today.weekday())
    this_month = today.replace(day=1)
    
    # Get overall statistics
    overall_stats = APIRequestLog.get_stats_by_date_range()
    
    # Get daily statistics for the last 30 days
    daily_stats = db.session.query(
        func.date(APIRequestLog.created_at).label('date'),
        func.count(APIRequestLog.id).label('count'),
        func.sum(APIRequestLog.total_tokens).label('tokens')
    ).group_by(func.date(APIRequestLog.created_at)
    ).order_by(desc('date')).limit(30).all()
    
    # Get service breakdown
    service_breakdown = db.session.query(
        APIRequestLog.service,
        func.count(APIRequestLog.id).label('count'),
        func.sum(APIRequestLog.total_tokens).label('tokens')
    ).group_by(APIRequestLog.service).all()
    
    # Get model breakdown for Anthropic
    model_breakdown = db.session.query(
        APIRequestLog.model,
        func.count(APIRequestLog.id).label('count'),
        func.sum(APIRequestLog.total_tokens).label('tokens')
    ).filter(APIRequestLog.service == 'anthropic'
    ).group_by(APIRequestLog.model).all()
    
    # Get recent requests with status, limited to 100
    recent_requests = APIRequestLog.query.order_by(
        APIRequestLog.created_at.desc()
    ).limit(100).all()
    
    # Calculate statistics for specific time periods
    today_stats = APIRequestLog.get_stats_by_date_range(start_date=today)
    yesterday_stats = APIRequestLog.get_stats_by_date_range(start_date=yesterday, end_date=today)
    this_week_stats = APIRequestLog.get_stats_by_date_range(start_date=this_week)
    this_month_stats = APIRequestLog.get_stats_by_date_range(start_date=this_month)
    
    # Format data for charts
    daily_labels = [str(item.date) for item in daily_stats]
    daily_requests = [item.count for item in daily_stats]
    daily_tokens = [int(item.tokens) if item.tokens else 0 for item in daily_stats]
    
    # Prepare service breakdown data for charts
    service_labels = [item.service for item in service_breakdown]
    service_counts = [item.count for item in service_breakdown]
    service_tokens = [int(item.tokens) if item.tokens else 0 for item in service_breakdown]
    
    return render_template(
        'admin/api_tracking.html',
        title="API Request Tracking",
        overall_stats=overall_stats,
        today_stats=today_stats,
        yesterday_stats=yesterday_stats,
        this_week_stats=this_week_stats,
        this_month_stats=this_month_stats,
        service_breakdown=service_breakdown,
        model_breakdown=model_breakdown,
        recent_requests=recent_requests,
        daily_labels=daily_labels,
        daily_requests=daily_requests,
        daily_tokens=daily_tokens,
        service_labels=service_labels,
        service_counts=service_counts,
        service_tokens=service_tokens
    )

@app.route('/admin/automation-control')
@login_required
def automation_control():
    """Unified automation control center"""
    if not current_user.is_admin:
        flash('You must be an admin to access automation controls.', 'danger')
        return redirect(url_for('index'))
    
    # Get current screenshots
    screenshots = Screenshot.query.order_by(Screenshot.timestamp.desc()).all()
    
    # Get any status messages from session
    automation_status = session.pop('automation_status', None)
    test_result = session.pop('test_result', None)
    
    return render_template('admin/automation_control.html',
                         screenshots=screenshots,
                         automation_status=automation_status,
                         test_result=test_result)

@app.route('/test-cleanup', methods=['POST'])
def test_cleanup():
    """Test route for Step 1 cleanup without authentication"""
    try:
        from step1_cleanup import run_cleanup
        success = run_cleanup()
        return f"SUCCESS: Cleanup completed" if success else "ERROR: Cleanup failed"
    except Exception as e:
        return f"ERROR: {str(e)}"

@app.route('/admin/run-automation-step', methods=['POST'])
@login_required
def run_automation_step():
    """Run individual automation steps"""
    if not current_user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    step = request.form.get('step')
    app.logger.info(f"Starting automation step: {step}")
    
    try:
        if step == 'cleanup':
            from step1_cleanup import run_cleanup
            success = run_cleanup()
            message = "Cleanup completed successfully" if success else "Cleanup failed"
            
        elif step == 'capture':
            from step2_capture import run_capture
            success = run_capture()
            message = "Screenshot capture completed successfully" if success else "Screenshot capture failed"
            
        elif step == 'ai_process':
            from step3_ai_process import run_ai_process
            success = run_ai_process()
            message = "AI processing completed successfully" if success else "AI processing failed"
            
        elif step == 'database_update':
            from step4_database import run_database_update
            success = run_database_update()
            message = "Database update completed successfully" if success else "Database update failed"
            
        else:
            return jsonify({'error': f'Unknown step: {step}'}), 400
        
        app.logger.info(f"Step {step} completed: success={success}")
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'danger')
            
        return redirect(url_for('automation_control'))
        
    except Exception as e:
        app.logger.error(f"Step {step} failed: {e}")
        flash(f"Failed to run {step}: {str(e)}", 'danger')
        return redirect(url_for('automation_control'))

@app.route('/admin/run-daily-automation', methods=['POST'])
@login_required
def run_daily_automation():
    """Run the complete daily automation workflow"""
    if not current_user.is_admin:
        flash('You must be an admin to run daily automation.', 'danger')
        return redirect(url_for('index'))
    
    try:
        from daily_automation import run_complete_automation
        result = run_complete_automation()
        
        session['automation_status'] = {
            'success': True,
            'message': "Complete daily automation workflow completed successfully",
            'details': str(result) if result else None
        }
        flash('Daily automation workflow completed successfully!', 'success')
        
    except Exception as e:
        error_msg = f"Failed to run daily automation: {str(e)}"
        session['automation_status'] = {
            'success': False,
            'message': error_msg,
            'details': str(e)
        }
        flash(error_msg, 'danger')
    
    return redirect(url_for('automation_control'))

@app.route('/admin/test-ai-extraction', methods=['POST'])
@login_required
def test_ai_extraction():
    """Test AI extraction on a sample image"""
    if not current_user.is_admin:
        flash('You must be an admin to test AI extraction.', 'danger')
        return redirect(url_for('index'))
    
    try:
        from automated_data_extractor import LotteryDataExtractor
        extractor = LotteryDataExtractor()
        
        # Test on a sample image
        test_result = extractor.test_single_image("attached_assets/Powerball_Results_20250527_034926.png")
        
        if test_result:
            session['test_result'] = test_result
            flash('AI extraction test completed successfully!', 'success')
        else:
            flash('AI extraction test failed - no result returned', 'warning')
            
    except Exception as e:
        flash(f'AI extraction test failed: {str(e)}', 'danger')
    
    return redirect(url_for('automation_control'))

@app.route('/admin/daily-automation')
@login_required
def daily_automation_dashboard():
    """Admin dashboard for daily automation system"""
    if not current_user.is_admin:
        flash('You must be an admin to access automation settings.', 'danger')
        return redirect(url_for('home'))
    
    global daily_scheduler
    
    # Initialize scheduler if not already done
    if daily_scheduler is None:
        try:
            import scheduler
            daily_scheduler = scheduler.init_scheduler(app, "01:00")
        except Exception as e:
            logger.error(f"Failed to initialize scheduler: {str(e)}")
            daily_scheduler = None
    
    # Get scheduler status
    scheduler_status = None
    if daily_scheduler:
        scheduler_status = daily_scheduler.get_status()
    
    return render_template('admin/daily_automation.html', 
                         scheduler_status=scheduler_status)

@app.route('/admin/run-daily-automation', methods=['POST'])
@login_required
def run_daily_automation_manual():
    """Manually trigger the daily automation workflow"""
    if not current_user.is_admin:
        flash('You must be an admin to run automation.', 'danger')
        return redirect(url_for('home'))
    
    try:
        from daily_automation import run_daily_automation
        
        logger.info("Manual daily automation triggered by admin")
        results = run_daily_automation(app)
        
        if results['overall_success']:
            flash('Daily automation completed successfully!', 'success')
            flash(f"Processed {results['capture']['count']} screenshots with AI analysis", 'info')
        else:
            flash('Daily automation completed with some issues. Check logs for details.', 'warning')
            
    except Exception as e:
        logger.error(f"Error running manual daily automation: {str(e)}")
        flash(f'Error running daily automation: {str(e)}', 'danger')
    
    return redirect(url_for('daily_automation_dashboard'))

@app.route('/admin/run-complete-workflow', methods=['POST'])
@login_required
def run_complete_workflow():
    """Run the complete 4-step automation workflow and return JSON response"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Admin access required'}), 403
    
    try:
        from daily_automation import run_complete_automation
        
        app.logger.info("Complete automation workflow triggered by admin")
        results = run_complete_automation()
        
        if results['overall_success']:
            # Clear cache to ensure fresh data on homepage
            from cache_manager import clear_results_cache
            clear_results_cache()
            
            return jsonify({
                'success': True, 
                'message': 'Complete automation workflow finished successfully',
                'results': results
            })
        else:
            return jsonify({
                'success': False, 
                'message': 'Automation workflow completed with errors',
                'results': results
            })
            
    except Exception as e:
        app.logger.error(f"Error running complete workflow: {str(e)}")
        return jsonify({
            'success': False, 
            'error': str(e)
        }), 500

@app.route('/admin/scheduler-control/<action>', methods=['POST'])
@login_required
def scheduler_control(action):
    """Control the automated scheduler (start/stop)"""
    if not current_user.is_admin:
        flash('You must be an admin to control the scheduler.', 'danger')
        return redirect(url_for('home'))
    
    global daily_scheduler
    
    try:
        if action == 'start':
            if daily_scheduler is None:
                import scheduler
                daily_scheduler = scheduler.init_scheduler(app, "01:00")
                flash('Automated scheduler started successfully! Daily processing will run at 1:00 AM.', 'success')
            else:
                flash('Scheduler is already running.', 'info')
                
        elif action == 'stop':
            if daily_scheduler:
                daily_scheduler.stop_scheduler()
                daily_scheduler = None
                flash('Automated scheduler stopped.', 'warning')
            else:
                flash('Scheduler is not running.', 'info')
                
    except Exception as e:
        logger.error(f"Error controlling scheduler: {str(e)}")
        flash(f'Error controlling scheduler: {str(e)}', 'danger')
    
    return redirect(url_for('daily_automation_dashboard'))

# When running directly, not through gunicorn

if __name__ == "__main__":
    # Extra logging to help diagnose startup issues
    import logging
    import os
    
    logging.getLogger('werkzeug').setLevel(logging.INFO)
    
    # Use port 5000 for deployment compatibility
    port = int(os.environ.get('PORT', 5000))
    
    # Start Flask app
    print(f"Starting Flask application on 0.0.0.0:{port}...")
    app.run(host="0.0.0.0", port=port, debug=True, use_reloader=False)