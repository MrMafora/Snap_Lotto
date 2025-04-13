"""
Application configuration and setup.

This file contains the application routes and business logic for the lottery application.
It provides a factory function to create the Flask application, which is then imported
in main.py for gunicorn deployment.
"""
import os
import json
import logging
import shutil
import traceback
from functools import wraps
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, send_file
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf import FlaskForm, CSRFProtect
from wtforms import StringField, PasswordField, SubmitField, BooleanField, EmailField
from wtforms.validators import DataRequired, Email, Length, EqualTo, ValidationError
from models import db, Screenshot, LotteryResult, ScheduleConfig, User
from data_aggregator import aggregate_data, validate_and_correct_known_draws
from scheduler import run_lottery_task
from ticket_scanner import process_ticket_image
from import_snap_lotto_data import import_snap_lotto_data

# Set up module-specific logger
from logger import setup_logger
logger = setup_logger(__name__, level=logging.DEBUG)

# Application factory function
def create_app():
    """Create and configure the Flask application"""
    app = Flask(__name__)
    app.secret_key = os.environ.get("SESSION_SECRET", "lottery-scraper-default-secret")
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
    
    # Configure database
    app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///lottery.db")
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "pool_recycle": 300,
        "pool_pre_ping": True
    }
    
    # Only add SSL requirement if DATABASE_URL is set and contains postgresql (production environment)
    if os.environ.get("DATABASE_URL") and 'postgresql' in os.environ.get("DATABASE_URL", ""):
        app.config["SQLALCHEMY_ENGINE_OPTIONS"]["connect_args"] = {"sslmode": "require"}
    
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    
    # Initialize the app with the database
    db.init_app(app)
    
    # Initialize Flask-Login
    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'Please log in to access this page'
    login_manager.login_message_category = 'info'
    
    # Initialize CSRF protection
    csrf = CSRFProtect(app)
    
    # User loader for Flask-Login
    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))
    
    # Create all database tables
    with app.app_context():
        db.create_all()
    
    # Import scheduler components
    from scheduler import init_scheduler, schedule_task, remove_task
    
    # Initialize scheduler
    scheduler = init_scheduler(app)
    
    # Setup initial schedules
    with app.app_context():
        # Run validation to correct any known draw issues
        try:
            corrected_count = validate_and_correct_known_draws()
            if corrected_count > 0:
                logger.info(f"Corrected {corrected_count} draws with known good data")
        except Exception as e:
            logger.error(f"Error validating known draws: {str(e)}")
            
        # Set up initial schedules if none exist
        if ScheduleConfig.query.count() == 0:
            # Add default lottery URLs
            default_urls = [
                {'url': 'https://www.nationallottery.co.za/lotto-history', 'lottery_type': 'Lotto'},
                {'url': 'https://www.nationallottery.co.za/lotto-plus-1-history', 'lottery_type': 'Lotto Plus 1'},
                {'url': 'https://www.nationallottery.co.za/lotto-plus-2-history', 'lottery_type': 'Lotto Plus 2'},
                {'url': 'https://www.nationallottery.co.za/powerball-history', 'lottery_type': 'Powerball'},
                {'url': 'https://www.nationallottery.co.za/powerball-plus-history', 'lottery_type': 'Powerball Plus'},
                {'url': 'https://www.nationallottery.co.za/daily-lotto-history', 'lottery_type': 'Daily Lotto'}
            ]
            
            # Add results pages
            results_urls = [
                {'url': 'https://www.nationallottery.co.za/results/lotto', 'lottery_type': 'Lotto Results'},
                {'url': 'https://www.nationallottery.co.za/results/lotto-plus-1-results', 'lottery_type': 'Lotto Plus 1 Results'},
                {'url': 'https://www.nationallottery.co.za/results/lotto-plus-2-results', 'lottery_type': 'Lotto Plus 2 Results'},
                {'url': 'https://www.nationallottery.co.za/results/powerball', 'lottery_type': 'Powerball Results'},
                {'url': 'https://www.nationallottery.co.za/results/powerball-plus', 'lottery_type': 'Powerball Plus Results'},
                {'url': 'https://www.nationallottery.co.za/results/daily-lotto', 'lottery_type': 'Daily Lotto Results'}
            ]
            
            # Combine all URLs
            all_urls = default_urls + results_urls
            
            for i, config in enumerate(all_urls):
                # Stagger the scheduled times
                hour = 1  # Run at 1 AM
                minute = i * 5  # 5 minutes apart
                
                schedule = ScheduleConfig(
                    url=config['url'],
                    lottery_type=config['lottery_type'],
                    frequency='daily',
                    hour=hour,
                    minute=minute,
                    active=True
                )
                db.session.add(schedule)
            
            db.session.commit()
            logger.info("Added default lottery schedule configurations")
        
        # Schedule active tasks
        for config in ScheduleConfig.query.filter_by(active=True).all():
            schedule_task(scheduler, config)
    
    # Form definitions
    class LoginForm(FlaskForm):
        """Login form for admin authentication"""
        username = StringField('Username', validators=[DataRequired()])
        password = PasswordField('Password', validators=[DataRequired()])
        remember = BooleanField('Remember Me')
        submit = SubmitField('Log In')

    class RegistrationForm(FlaskForm):
        """Registration form for admin accounts"""
        username = StringField('Username', validators=[DataRequired(), Length(min=3, max=50)])
        email = EmailField('Email', validators=[DataRequired(), Email()])
        password = PasswordField('Password', validators=[DataRequired(), Length(min=8)])
        confirm_password = PasswordField('Confirm Password', validators=[DataRequired(), EqualTo('password')])
        submit = SubmitField('Create Account')
        
        # Custom validators
        def validate_username(self, username):
            user = User.query.filter_by(username=username.data).first()
            if user:
                raise ValidationError('Username is already taken. Please choose another one.')
        
        def validate_email(self, email):
            user = User.query.filter_by(email=email.data).first()
            if user:
                raise ValidationError('Email is already registered. Please use another one or reset your password.')

    # Admin access decorator
    def admin_required(f):
        """Decorator for routes that require admin privileges"""
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or not current_user.is_admin:
                flash("You don't have permission to access this page. Please log in as an administrator.", "warning")
                return redirect(url_for('login', next=request.url))
            return f(*args, **kwargs)
        return decorated_function

    # Register routes
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        """Admin login route"""
        if current_user.is_authenticated:
            return redirect(url_for('index'))
        
        form = LoginForm()
        if form.validate_on_submit():
            user = User.query.filter_by(username=form.username.data).first()
            if user and user.check_password(form.password.data):
                login_user(user, remember=form.remember.data)
                next_page = request.args.get('next')
                return redirect(next_page if next_page else url_for('admin_dashboard'))
            else:
                flash('Login failed. Please check your username and password.', 'danger')
        
        return render_template('login.html', form=form)

    @app.route('/logout')
    @login_required
    def logout():
        """Admin logout route"""
        logout_user()
        flash('You have been logged out.', 'info')
        return redirect(url_for('index'))

    @app.route('/admin/register', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def register():
        """Admin registration route (only accessible by existing admins)"""
        form = RegistrationForm()
        if form.validate_on_submit():
            user = User(
                username=form.username.data,
                email=form.email.data,
                is_admin=True
            )
            user.set_password(form.password.data)
            db.session.add(user)
            db.session.commit()
            flash(f'Account created for {form.username.data}!', 'success')
            return redirect(url_for('admin_dashboard'))
        
        return render_template('register.html', form=form)

    @app.route('/admin')
    @login_required
    @admin_required
    def admin_dashboard():
        """Admin dashboard showing links to admin-only features"""
        return render_template('admin/dashboard.html')

    @app.route('/')
    def index():
        """Home page showing lottery results and ticket scanning feature"""
        latest_results = LotteryResult.query.order_by(LotteryResult.draw_date.desc()).limit(10).all()
        return render_template('index.html', results=latest_results)

    @app.route('/results')
    def results():
        """Page showing all lottery results"""
        # Run validation to ensure known correct data is used
        try:
            corrected = validate_and_correct_known_draws()
            if corrected > 0:
                logger.info(f"Corrected {corrected} draws with known correct data")
        except Exception as e:
            logger.error(f"Error validating known draws: {str(e)}")
        
        lottery_type = request.args.get('lottery_type', None)
        page = request.args.get('page', 1, type=int)
        per_page = 20
        
        # Import the normalization functions
        from data_aggregator import normalize_lottery_type
        
        # Create a query that avoids showing duplicate entries
        query = db.session.query(LotteryResult)
        
        # Filter by lottery type if specified
        if lottery_type:
            # Remove "Results" suffix if present for consistent filtering
            normalized_type = normalize_lottery_type(lottery_type)
            query = query.filter(
                # Match either exact type or the same type with "Results" suffix
                db.or_(
                    LotteryResult.lottery_type == normalized_type,
                    LotteryResult.lottery_type == f"{normalized_type} Results"
                )
            )
        
        # Get all results to remove duplicates properly
        all_results = query.order_by(LotteryResult.draw_date.desc()).all()
        
        # Process results to remove duplicates
        unique_results = []
        processed_draws = set()
        
        for result in all_results:
            # Normalize draw number to handle both "2530" and "LOTTO DRAW 2530" formats
            from data_aggregator import normalize_draw_number
            draw_key = normalize_draw_number(result.draw_number)
            normalized_type = normalize_lottery_type(result.lottery_type)
            
            # Create a unique key for this draw
            unique_key = f"{normalized_type}_{draw_key}"
            
            # Skip if we've already processed this draw
            if unique_key in processed_draws:
                continue
            
            # Mark this draw as processed
            processed_draws.add(unique_key)
            
            # Add to unique results
            unique_results.append(result)
        
        # Implement manual pagination since we're doing in-memory filtering
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paged_results = unique_results[start_idx:end_idx]
        
        # Create a pagination class similar to Flask-SQLAlchemy's but with added iter_pages method
        class Pagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page  # Ceiling division
                self.has_next = page < self.pages
                self.has_prev = page > 1
                
            @property
            def prev_num(self):
                return self.page - 1 if self.has_prev else None
                
            @property
            def next_num(self):
                return self.page + 1 if self.has_next else None
                
            def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
                """Iterate over page numbers."""
                last = 0
                for num in range(1, self.pages + 1):
                    if (num <= left_edge or
                        (self.page - left_current - 1 < num < self.page + right_current) or
                        num > self.pages - right_edge):
                        if last + 1 != num:
                            yield None
                        yield num
                        last = num
        
        # Create our pagination object
        total = len(unique_results)
        pagination = Pagination(
            items=paged_results,
            page=page,
            per_page=per_page,
            total=total
        )
        
        return render_template('results.html', results=pagination, lottery_type=lottery_type)
    
    @app.route('/ticket-scanner')
    def ticket_scanner():
        """Page for scanning lottery tickets and checking if they won"""
        class EmptyForm(FlaskForm):
            """Empty form for CSRF protection"""
            pass
            
        form = EmptyForm()
        return render_template('ticket_scanner_new.html', form=form)
    
    @app.route('/scan-ticket', methods=['POST'])
    def scan_ticket():
        """API endpoint to process a lottery ticket image and check if it's a winner"""
        logger.info("Scan ticket endpoint called")
        
        try:
            # Get lottery type and draw number from form if provided
            lottery_type = request.form.get('lottery_type', '')
            draw_number = request.form.get('draw_number', '')
            
            # Get image from request
            if 'ticket_image' not in request.files:
                logger.error("No image file in request")
                return jsonify({'error': 'No image file provided'}), 400
                
            image_file = request.files['ticket_image']
            if image_file.filename == '':
                logger.error("Empty filename")
                return jsonify({'error': 'No image selected'}), 400
                
            # Process the image and check for winning numbers
            logger.info(f"Processing ticket image: {image_file.filename}")
            logger.info(f"Lottery type: {lottery_type}")
            logger.info(f"Draw number: {draw_number}")
            
            # Read image file
            image_data = image_file.read()
            logger.info(f"Image data read successfully: {len(image_data)} bytes")
            
            # Detect file extension from filename
            file_ext = os.path.splitext(image_file.filename)[1].lower()
            logger.info(f"Detected file extension: {file_ext}")
            
            # Call the ticket scanner function
            logger.info("Calling process_ticket_image function")
            result = process_ticket_image(image_data, lottery_type, draw_number, file_ext)
            
            if 'error' in result:
                return jsonify(result), 400
                
            logger.info(f"Ticket processing result: {json.dumps(result)}")
            # Add show_ad flag to signal the frontend to show the ad overlay
            result['show_ad'] = True
            result['ad_type'] = 'interstitial'
            return jsonify(result)
            
        except Exception as e:
            logger.exception(f"Error processing ticket: {str(e)}")
            return jsonify({'error': f'Error processing ticket: {str(e)}'}), 500

    @app.route('/settings', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def settings():
        """Page for configuring screenshot schedules (admin only)"""
        if request.method == 'POST':
            url = request.form.get('url')
            lottery_type = request.form.get('lottery_type')
            frequency = request.form.get('frequency', 'daily')
            hour = request.form.get('hour', 0, type=int)
            minute = request.form.get('minute', 0, type=int)
            
            if not url or not lottery_type:
                flash('URL and lottery type are required', 'danger')
                return redirect(url_for('settings'))
            
            # Create or update schedule
            config = ScheduleConfig.query.filter_by(url=url).first()
            if not config:
                config = ScheduleConfig(url=url, lottery_type=lottery_type, 
                                      frequency=frequency, hour=hour, minute=minute, active=True)
                db.session.add(config)
            else:
                config.lottery_type = lottery_type
                config.frequency = frequency
                config.hour = hour
                config.minute = minute
                config.active = True
            
            db.session.commit()
            
            # Add to scheduler
            schedule_task(scheduler, config)
            
            flash('Schedule updated successfully', 'success')
            return redirect(url_for('settings'))
        
        schedules = ScheduleConfig.query.all()
        return render_template('settings.html', schedules=schedules)
        
    @app.route('/retake_screenshots', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def retake_screenshots():
        """Retake screenshots for all configured URLs"""
        from scheduler import retake_all_screenshots
        
        if request.method == 'POST':
            # Run the screenshot retake process
            try:
                results = retake_all_screenshots()
                
                # Count successes and failures
                success_count = sum(1 for result in results.values() if isinstance(result, dict) and result.get('status') == 'success')
                error_count = len(results) - success_count if isinstance(results, dict) else 0
                
                if success_count > 0:
                    flash(f"Successfully retook {success_count} screenshots. {error_count} screenshots failed.", "success")
                else:
                    flash(f"Failed to retake screenshots. Please check the logs for details.", "danger")
                    
            except Exception as e:
                logger.error(f"Error retaking screenshots: {str(e)}")
                flash(f"Error retaking screenshots: {str(e)}", "danger")
                
            return redirect(url_for('settings'))
        else:
            # Render confirmation page
            return render_template('retake_screenshots.html')
    
    @app.route('/toggle_schedule/<int:id>')
    @login_required
    @admin_required
    def toggle_schedule(id):
        """Toggle a schedule on or off (admin only)"""
        config = ScheduleConfig.query.get_or_404(id)
        config.active = not config.active
        db.session.commit()
        
        if config.active:
            schedule_task(scheduler, config)
            flash(f'Schedule for {config.lottery_type} activated', 'success')
        else:
            remove_task(scheduler, config.id)
            flash(f'Schedule for {config.lottery_type} deactivated', 'warning')
            
        return redirect(url_for('settings'))
        
    @app.route('/delete_schedule/<int:id>')
    @login_required
    @admin_required
    def delete_schedule(id):
        """Delete a schedule (admin only)"""
        config = ScheduleConfig.query.get_or_404(id)
        remove_task(scheduler, config.id)
        db.session.delete(config)
        db.session.commit()
        flash(f'Schedule for {config.lottery_type} deleted', 'warning')
        return redirect(url_for('settings'))
        
    @app.route('/api/run_now/<int:id>')
    @login_required
    @admin_required
    def run_now(id):
        """Manually run a scheduled task immediately in a background thread (admin only)"""
        config = ScheduleConfig.query.get_or_404(id)
        
        # Simply call the task function directly and let it handle threading
        logger.info(f"Manually running task for {config.url}")
        run_lottery_task(config.url, config.lottery_type)
        
        # Return immediately
        return jsonify({
            'status': 'success', 
            'message': f'Data sync started for {config.lottery_type}. This may take 30-60 seconds to complete.'
        })
        
    @app.route('/api/screenshots')
    def get_screenshots():
        """API endpoint to fetch recent screenshots"""
        screenshots = Screenshot.query.order_by(Screenshot.timestamp.desc()).limit(20).all()
        return jsonify([{
            'id': s.id,
            'url': s.url,
            'lottery_type': s.lottery_type,
            'timestamp': s.timestamp.isoformat(),
            'path': s.path
        } for s in screenshots])
        
    @app.route('/api/results/<lottery_type>')
    def get_results(lottery_type):
        """API endpoint to fetch results for a specific lottery type"""
        limit = request.args.get('limit', 5, type=int)
        
        # Import normalization functions
        from data_aggregator import normalize_lottery_type
        
        # Normalize the lottery type to handle suffix differences
        normalized_type = normalize_lottery_type(lottery_type)
        
        # Query for results, handling both exact and suffix matches
        results = LotteryResult.query.filter(
            db.or_(
                LotteryResult.lottery_type == normalized_type,
                LotteryResult.lottery_type == f"{normalized_type} Results"
            )
        ).order_by(LotteryResult.draw_date.desc()).limit(limit).all()
        
        return jsonify([result.to_dict() for result in results])
        
    @app.route('/visualizations')
    def visualizations():
        """Data visualization dashboard for lottery results"""
        return render_template('visualizations.html')
        
    @app.route('/import-data', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def import_data():
        """Page for importing lottery data from spreadsheets (admin only)"""
        imported_results = []
        messages = []
        
        if request.method == 'POST':
            # Check if a file was uploaded
            if 'excel_file' not in request.files:
                messages.append(('danger', 'No file part in the request. Please select a file.'))
                return render_template('import.html', messages=messages)
                
            file = request.files['excel_file']
            
            # If user doesn't select file, browser submits an empty part without filename
            if file.filename == '':
                messages.append(('danger', 'No selected file. Please choose an Excel file to upload.'))
                return render_template('import.html', messages=messages)
                
            # Check for valid Excel file
            if file and file.filename.endswith(('.xlsx', '.xls')):
                # Create uploads directory if it doesn't exist
                uploads_dir = os.path.join(os.getcwd(), 'uploads')
                os.makedirs(uploads_dir, exist_ok=True)
                
                # Save the file
                filename = secure_filename(file.filename)
                file_path = os.path.join(uploads_dir, filename)
                file.save(file_path)
                
                # Process the uploaded file
                # Determine import type
                import_type = request.form.get('import_type', 'standard')
                logger.info(f"Importing file: {filename} as type: {import_type}")
                
                try:
                    if import_type == 'snap_lotto':
                        # Use the specialized importer for Snap Lotto format
                        logger.info("Using Snap Lotto importer")
                        success = import_snap_lotto_data(file_path, flask_app=app)
                        if success:
                            messages.append(('success', f'Successfully imported Snap Lotto data from {filename}'))
                            # Get recently imported results
                            imported_results = LotteryResult.query.order_by(LotteryResult.created_at.desc()).limit(50).all()
                        else:
                            messages.append(('danger', f'Failed to import Snap Lotto data from {filename}'))
                    else:
                        # Use the standard importer
                        logger.info("Using standard Excel importer")
                        from import_excel import import_excel_data
                        import_excel_data(file_path)
                        messages.append(('success', f'Successfully imported data from {filename}'))
                        # Get recently imported results
                        imported_results = LotteryResult.query.order_by(LotteryResult.created_at.desc()).limit(50).all()
                except Exception as e:
                    logger.error(f"Error importing data: {str(e)}", exc_info=True)
                    messages.append(('danger', f'Error importing data: {str(e)}'))
                
                # Clean up the uploaded file
                try:
                    os.remove(file_path)
                    logger.info(f"Removed temporary file: {file_path}")
                except Exception as e:
                    logger.error(f"Error removing temporary file: {str(e)}")
            else:
                messages.append(('danger', 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'))
                
        return render_template('import.html', messages=messages, imported_results=imported_results)
        
    @app.route('/api/visualization-data')
    def visualization_data():
        """API endpoint to fetch data for visualizations"""
        # Import the normalization functions
        from data_aggregator import normalize_lottery_type
        
        lottery_type = request.args.get('lottery_type', None)
        data_type = request.args.get('data_type', 'numbers_frequency')
        
        if lottery_type:
            normalized_type = normalize_lottery_type(lottery_type)
            query = LotteryResult.query.filter(
                db.or_(
                    LotteryResult.lottery_type == normalized_type,
                    LotteryResult.lottery_type == f"{normalized_type} Results"
                )
            )
        else:
            query = LotteryResult.query
        
        results = query.order_by(LotteryResult.draw_date.desc()).limit(100).all()
        
        # Check if we have any results to work with
        if not results:
            return jsonify({
                'error': 'No data available',
                'message': 'No lottery results found for the selected filters.'
            })
        
        if data_type == 'numbers_frequency':
            # Calculate frequency of each number
            frequencies = {}
            total_draws = len(results)
            
            # Initialize all possible numbers based on lottery type
            max_number = 49  # Default for regular lotto
            if lottery_type and 'daily' in lottery_type.lower():
                max_number = 36  # Daily Lotto uses 1-36
            
            # Initialize all possible numbers with zero frequency
            for i in range(1, max_number + 1):
                frequencies[i] = 0
                
            # Count actual frequencies
            for result in results:
                numbers = result.get_numbers_list()
                for num in numbers:
                    if num in frequencies:
                        frequencies[num] += 1
            
            # Convert to sorted list of dictionaries for easier manipulation in JavaScript
            sorted_frequencies = []
            for num, freq in sorted(frequencies.items(), key=lambda x: int(x[0])):
                sorted_frequencies.append({
                    'number': str(num),
                    'frequency': freq
                })
            
            # Format for Chart.js
            chart_data = {
                'labels': [item['number'] for item in sorted_frequencies],
                'datasets': [{
                    'label': f'Number Frequency for {lottery_type if lottery_type else "All Types"}',
                    'data': [item['frequency'] for item in sorted_frequencies],
                    'backgroundColor': 'rgba(54, 162, 235, 0.6)',
                    'borderColor': 'rgba(54, 162, 235, 1)',
                    'borderWidth': 1
                }],
                'frequencies': sorted_frequencies,  # Add sorted frequency data for better insights
                'totalDraws': total_draws
            }
            
            return jsonify(chart_data)
        
        # Default response if no recognized data type
        return jsonify({
            'error': 'Invalid data type',
            'message': f'The data type {data_type} is not supported'
        })
    
    @app.route('/draw/<lottery_type>/<draw_number>')
    def draw_details(lottery_type, draw_number):
        """Page for viewing detailed prize payout information for a specific draw"""
        # Import normalization functions
        from data_aggregator import normalize_lottery_type, normalize_draw_number
        
        # Normalize inputs for consistent lookup
        normalized_type = normalize_lottery_type(lottery_type)
        normalized_draw = normalize_draw_number(draw_number)
        
        # Find the result
        result = LotteryResult.query.filter(
            db.and_(
                db.or_(
                    LotteryResult.lottery_type == normalized_type,
                    LotteryResult.lottery_type == f"{normalized_type} Results"
                ),
                db.func.lower(LotteryResult.draw_number).contains(normalized_draw.lower())
            )
        ).first()
        
        if not result:
            flash(f"Draw {draw_number} for {lottery_type} not found", "warning")
            return redirect(url_for('results'))
        
        return render_template('draw_details.html', result=result)
    
    @app.route('/admin/export-screenshots')
    @login_required
    @admin_required
    def export_screenshots():
        """Export latest screenshots (one per URL) and an empty template Excel file as a zip archive"""
        import io
        import zipfile
        import openpyxl
        from datetime import datetime
        from sqlalchemy import func
        
        # Create an in-memory file
        memory_file = io.BytesIO()
        
        # Create a zipfile in memory
        with zipfile.ZipFile(memory_file, 'w') as zf:
            # Get the latest screenshot for each unique URL
            # Using a subquery to get the latest screenshot ID for each URL
            subquery = db.session.query(
                Screenshot.url,
                func.max(Screenshot.timestamp).label('max_timestamp')
            ).group_by(Screenshot.url).subquery()
            
            # Join with the main table to get the actual screenshot records
            latest_screenshots = db.session.query(Screenshot).join(
                subquery,
                db.and_(
                    Screenshot.url == subquery.c.url,
                    Screenshot.timestamp == subquery.c.max_timestamp
                )
            ).all()
            
            # Create a manifest file with details of the screenshots
            manifest_content = "Screenshot Export Manifest\n"
            manifest_content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            manifest_content += f"Total Screenshots: {len(latest_screenshots)}\n\n"
            manifest_content += "Screenshot Details:\n"
            
            # Add screenshots to the zip file
            for idx, screenshot in enumerate(latest_screenshots):
                try:
                    # Get original screenshot path
                    screenshot_path = screenshot.path
                    
                    # Add to manifest
                    manifest_content += f"{idx+1}. {screenshot.lottery_type} - {screenshot.timestamp.strftime('%Y-%m-%d %H:%M:%S')}\n"
                    manifest_content += f"   URL: {screenshot.url}\n"
                    
                    # Only add the file if it exists
                    if os.path.exists(screenshot_path):
                        # Add to zip with a nicely formatted filename
                        formatted_name = f"{screenshot.lottery_type.replace(' ', '_')}_{screenshot.timestamp.strftime('%Y%m%d_%H%M%S')}.png"
                        zf.write(screenshot_path, f"screenshots/{formatted_name}")
                        
                        # Also add zoomed screenshot if available
                        if screenshot.zoomed_path and os.path.exists(screenshot.zoomed_path):
                            formatted_zoomed_name = f"{screenshot.lottery_type.replace(' ', '_')}_{screenshot.timestamp.strftime('%Y%m%d_%H%M%S')}_zoomed.png"
                            zf.write(screenshot.zoomed_path, f"screenshots/{formatted_zoomed_name}")
                    else:
                        manifest_content += f"   WARNING: Screenshot file not found: {screenshot_path}\n"
                except Exception as e:
                    logger.error(f"Error adding screenshot to zip: {str(e)}")
                    manifest_content += f"   ERROR: Failed to add screenshot: {str(e)}\n"
            
            # Add manifest to zip
            zf.writestr("manifest.txt", manifest_content)
            
            # Create an empty Excel template
            workbook = openpyxl.Workbook()
            
            # Create sheet for each lottery type
            lottery_types = ["Lotto", "Lotto Plus 1", "Lotto Plus 2", "Powerball", "Powerball Plus", "Daily Lotto"]
            
            # Delete default sheet
            if "Sheet" in workbook.sheetnames:
                del workbook[workbook.sheetnames[0]]
            
            for lottery_type in lottery_types:
                sheet = workbook.create_sheet(title=lottery_type)
                
                # Add headers
                headers = ["Draw Number", "Draw Date", "Number 1", "Number 2", "Number 3", "Number 4", "Number 5"]
                
                # Add bonus ball header for games that have it
                if lottery_type != "Daily Lotto":
                    if "Powerball" in lottery_type:
                        headers.append("Powerball")
                    else:
                        headers.append("Bonus Ball")
                        
                # Add division headers - typically there are up to 8 divisions
                for i in range(1, 9):
                    headers.append(f"Division {i} Winners")
                    headers.append(f"Division {i} Prize")
                
                # Set headers
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
                
                # Format header row
                for cell in sheet[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
            
            # Create a usage instructions sheet
            instructions = workbook.create_sheet(title="Instructions")
            instructions['A1'] = "Lottery Data Import Template"
            instructions['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            
            instructions['A3'] = "Instructions:"
            instructions['A3'].font = openpyxl.styles.Font(bold=True)
            
            instructions_text = [
                "1. Enter lottery data in the corresponding sheets.",
                "2. For each draw, provide the draw number, date, and winning numbers.",
                "3. For division data, enter the number of winners and prize amount.",
                "4. Date format should be YYYY-MM-DD (e.g., 2025-04-13).",
                "5. Save this file when complete.",
                "6. Upload the completed file through the 'Import Data' page."
            ]
            
            for i, text in enumerate(instructions_text, 4):
                instructions[f'A{i}'] = text
            
            # Save Excel to a bytes buffer
            excel_bytes = io.BytesIO()
            workbook.save(excel_bytes)
            excel_bytes.seek(0)
            
            # Add Excel template to zip
            zf.writestr("lottery_data_template.xlsx", excel_bytes.getvalue())
        
        # Reset file pointer
        memory_file.seek(0)
        
        # Generate timestamp for filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Send the in-memory file as a response for the browser to download
        return send_file(
            memory_file, 
            mimetype='application/zip', 
            as_attachment=True,
            download_name=f'lottery_screenshots_export_{timestamp}.zip'
        )

    return app