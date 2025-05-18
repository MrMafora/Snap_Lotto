#!/usr/bin/env python3
"""
Script to create a test data file with realistic lottery data.
"""
import os
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_test_data_file(output_path):
    """Create a test data Excel file with realistic lottery data"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create a sheet for each lottery type
    lottery_types = [
        "Lottery", 
        "Lottery Plus 1", 
        "Lottery Plus 2", 
        "Powerball", 
        "Powerball Plus", 
        "Daily Lottery"
    ]
    
    for lottery_type in lottery_types:
        sheet = wb.create_sheet(title=lottery_type)
        
        # Set up headers
        headers = ["Draw Number", "Draw Date", "Numbers"]
        
        # Add bonus/powerball column for certain lottery types
        if lottery_type in ["Powerball", "Powerball Plus"]:
            headers.append("Powerball")
        
        # Add division columns for all types
        for i in range(1, 7):
            if lottery_type == "Daily Lottery" and i > 4:
                # Daily Lottery only has 4 divisions
                continue
            headers.append(f"Division {i} Winners")
            headers.append(f"Division {i} Payout")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            sheet.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Generate test data
        num_rows = random.randint(15, 25)  # Random number of draws
        
        # Determine max numbers and range based on lottery type
        if lottery_type == "Daily Lottery":
            main_numbers = 5
            main_range = 36
            has_bonus = False
        elif "Powerball" in lottery_type:
            main_numbers = 5
            main_range = 50
            has_bonus = True
            bonus_range = 20
        else:  # Lottery types
            main_numbers = 6
            main_range = 52
            has_bonus = False
        
        # Current date as starting point
        current_date = datetime.now()
        
        for row_num in range(2, num_rows + 2):
            col = 1
            
            # Draw number (decreases as we go back in time)
            draw_number = 2000 - (row_num - 2)
            sheet.cell(row=row_num, column=col, value=draw_number)
            col += 1
            
            # Draw date (goes back in time)
            draw_date = current_date - timedelta(days=(row_num - 2) * 7)
            sheet.cell(row=row_num, column=col, value=draw_date)
            col += 1
            
            # Generate winning numbers
            numbers = sorted(random.sample(range(1, main_range + 1), main_numbers))
            sheet.cell(row=row_num, column=col, value=", ".join(str(n) for n in numbers))
            col += 1
            
            # Bonus/Powerball for applicable types
            if has_bonus:
                bonus = random.randint(1, bonus_range)
                sheet.cell(row=row_num, column=col, value=bonus)
                col += 1
            
            # Division winners and payouts
            max_divisions = 4 if lottery_type == "Daily Lottery" else 6
            
            for div in range(1, max_divisions + 1):
                # Winners (higher divisions have fewer winners)
                winners = max(0, int(random.triangular(0, 100 / div, 10 / div)))
                sheet.cell(row=row_num, column=col, value=winners)
                col += 1
                
                # Payout (higher divisions have bigger payouts)
                base_payout = 1000000 / div if div == 1 else 10000 / (div - 1)
                variation = random.uniform(0.8, 1.2)
                payout = base_payout * variation
                
                # Format as currency with commas
                payout_str = f"R {payout:,.2f}"
                sheet.cell(row=row_num, column=col, value=payout_str)
                col += 1
    
    # Save the workbook
    wb.save(output_path)
    print(f"Test data created and saved to {output_path}")
    return True

if __name__ == "__main__":
    output_file = "lottery_test_data.xlsx"
    create_test_data_file(output_file)